#最初にxlsxファイルを取り込み、弁当注文人数の列を探して、0より大きい行だけ抽出して結合するコード
#"C:\Users\owner\Desktop\works\保管書類\店舗リスト\table_HHHL共通店舗一覧m.xlsm"をマスタとして参照している
#出力は「弁当チェック_年月_yyyymmdd_hhmm.xlsx」という名前で、ダウンロードフォルダに保存される
#出力ファイルのA列には、マスタにURLがあればハイパーリンクで表示される
#店舗名がマスタにあるがファイルが存在しない店舗は、末尾のセクションに「【ファイルなし店舗】レク費urlリストにあるがファイルが存在しない店舗」としてまとめて表示される（ただし、店舗名が類似しているものは除外される）
#店舗名が空欄のファイルは、ファイル名から店舗名を推測して照合する（例：ファイル名に「【○○店】」のように店舗名が含まれていれば、それを店舗名として扱う）
import pandas as pd
import openpyxl
import os
import re
import webbrowser
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox, ttk
import sys
import unicodedata
import warnings
import ctypes
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment

warnings.simplefilter('ignore', UserWarning)


def minimize_console_window():
    """起動時に表示される黒いコンソールウインドウを最小化する。
    pythonw実行やexe化時はコンソールが存在しないため、その場合は何もしない。"""
    try:
        hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        if hwnd:
            ctypes.windll.user32.ShowWindow(hwnd, 6)  # SW_MINIMIZE
    except Exception:
        pass

REF_PATH = r"C:\Users\owner\Desktop\works\保管書類\店舗リスト\table_HHHL共通店舗一覧m.xlsm"

# ============================== 仕事の進め方 / 取説 ==============================

INSTRUCTIONS_FILENAME = "手順.md"

DEFAULT_INSTRUCTIONS = """\
※この手順は変更されることがあります。内容を直接書き換えたい場合は
　右上の「✎ 取説を編集」ボタンを押してください(メモ帳でこのファイルが開きます)。
※URL(https://...)やファイルのパス(例: C:\\folder\\file.xlsx)をそのまま1行に
　書いておくと、「仕事の進め方」画面でクリックできるリンクになります。
"""

_LINK_CHAR = r"[^\s。、，」』（）\"']"
LINK_PATTERN = re.compile(rf"(https?://{_LINK_CHAR}+|[A-Za-z]:\\{_LINK_CHAR}+|\\\\{_LINK_CHAR}+)")
LINK_TRAILING_CHARS = "。、,.)）」』\"'"


def get_instructions_path():
    app_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(app_dir, INSTRUCTIONS_FILENAME)


def ensure_instructions_file():
    path = get_instructions_path()
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            f.write(DEFAULT_INSTRUCTIONS)
    return path


def read_instructions():
    path = ensure_instructions_file()
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except OSError as e:
        return f"手順ファイルの読み込みに失敗しました。\n{path}\n{e}"


def open_instructions_editor():
    path = ensure_instructions_file()
    import subprocess
    try:
        subprocess.Popen(["notepad.exe", path])
    except OSError as e:
        messagebox.showerror("編集に失敗しました", f"{path}\n{e}")


def open_link(href):
    if href.startswith("http://") or href.startswith("https://"):
        webbrowser.open(href)
        return
    try:
        os.startfile(href)
    except OSError as e:
        messagebox.showerror("開けませんでした", f"{href}\n{e}")


_MD_INLINE_PATTERN = re.compile(
    r"(?P<bold>\*\*(?P<boldtext>.+?)\*\*)"
    rf"|(?P<code>`(?P<codetext>[^`\n]+)`)"
    rf"|(?P<link>https?://{_LINK_CHAR}+|[A-Za-z]:\\{_LINK_CHAR}+|\\\\{_LINK_CHAR}+)"
)


def insert_instructions_with_links(text_widget, content):
    """Markdownの簡易記法(見出し・引用・コードブロック・太字・区切り線)をそれらしく
    整形しつつ、URL・ファイルパスをクリック可能なリンクとして挿入する。"""
    from tkinter import font as tkfont
    try:
        base_font = tkfont.nametofont(text_widget.cget("font"))
        family, size = base_font.actual("family"), base_font.actual("size")
    except Exception:
        family, size = "TkDefaultFont", 10

    text_widget.tag_configure("hyperlink", foreground="#2f6f63", underline=True)
    text_widget.tag_configure("md_h1", font=(family, size + 5, "bold"), spacing1=10, spacing3=6)
    text_widget.tag_configure("md_h2", font=(family, size + 3, "bold"), spacing1=8, spacing3=4)
    text_widget.tag_configure("md_h3", font=(family, size + 1, "bold"), spacing1=6, spacing3=3)
    text_widget.tag_configure("md_bold", font=(family, size, "bold"))
    text_widget.tag_configure("md_code", font=("Consolas", size), background="#eeeeee")
    text_widget.tag_configure("md_code_block", font=("Consolas", size), background="#f2f2f2", lmargin1=14, lmargin2=14)
    text_widget.tag_configure("md_quote", foreground="#4a4a4a", background="#f0efe6", lmargin1=20, lmargin2=20, spacing1=3, spacing3=3)
    text_widget.tag_configure("md_hr", foreground="#aaaaaa")

    orig_cursor = text_widget.cget("cursor")
    link_index = 0

    def insert_inline(text_line, block_tags):
        nonlocal link_index
        pos = 0
        for m in _MD_INLINE_PATTERN.finditer(text_line):
            start, end = m.span()
            if start > pos:
                text_widget.insert("end", text_line[pos:start], block_tags)
            if m.group("bold"):
                text_widget.insert("end", m.group("boldtext"), block_tags + ("md_bold",))
            elif m.group("code"):
                text_widget.insert("end", m.group("codetext"), block_tags + ("md_code",))
            else:
                raw = m.group("link")
                href = raw.rstrip(LINK_TRAILING_CHARS)
                trailing = raw[len(href):]
                tag_name = f"link_{link_index}"
                link_index += 1
                text_widget.insert("end", href, block_tags + ("hyperlink", tag_name))
                text_widget.tag_bind(tag_name, "<Button-1>", lambda e, href=href: open_link(href))
                text_widget.tag_bind(tag_name, "<Enter>", lambda e: text_widget.config(cursor="hand2"))
                text_widget.tag_bind(tag_name, "<Leave>", lambda e: text_widget.config(cursor=orig_cursor))
                if trailing:
                    text_widget.insert("end", trailing, block_tags)
            pos = end
        text_widget.insert("end", text_line[pos:], block_tags)

    in_code_block = False
    for raw_line in content.splitlines():
        stripped = raw_line.strip()

        if stripped.startswith("```"):
            in_code_block = not in_code_block
            continue
        if in_code_block:
            text_widget.insert("end", raw_line + "\n", ("md_code_block",))
            continue

        if re.fullmatch(r"[-_*]{3,}", stripped):
            text_widget.insert("end", "─" * 42 + "\n", ("md_hr",))
            continue

        m = re.match(r"^(#{1,6})\s+(.*)$", raw_line)
        if m:
            level = len(m.group(1))
            tag = "md_h1" if level == 1 else "md_h2" if level == 2 else "md_h3"
            insert_inline(m.group(2), (tag,))
            text_widget.insert("end", "\n")
            continue

        m = re.match(r"^>\s?(.*)$", raw_line)
        if m:
            inner = m.group(1)
            hm = re.match(r"^(#{1,6})\s+(.*)$", inner)
            if hm:
                level = len(hm.group(1))
                tag = "md_h1" if level == 1 else "md_h2" if level == 2 else "md_h3"
                insert_inline(hm.group(2), ("md_quote", tag))
            else:
                insert_inline(inner, ("md_quote",))
            text_widget.insert("end", "\n")
            continue

        insert_inline(raw_line, ())
        text_widget.insert("end", "\n")


def show_instructions_dialog():
    """起動直後に「仕事の進め方」を表示し、閉じるまで先へ進ませない。"""
    root = tk.Tk()
    root.withdraw()

    win = tk.Toplevel(root)
    win.title(f"仕事の進め方({os.path.basename(os.path.dirname(os.path.abspath(__file__)))})")
    win.geometry("560x520")

    text_frame = ttk.Frame(win)
    text_frame.pack(fill="both", expand=True, padx=14, pady=(14, 6))

    text_widget = tk.Text(text_frame, wrap="word")
    scroll = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
    text_widget.config(yscrollcommand=scroll.set)
    scroll.pack(side="right", fill="y")
    text_widget.pack(side="left", fill="both", expand=True)

    insert_instructions_with_links(text_widget, read_instructions())
    text_widget.config(state="disabled")

    button_row = ttk.Frame(win)
    button_row.pack(side="bottom", fill="x", padx=14, pady=(0, 14))
    ttk.Button(button_row, text="✎ 取説を編集", command=open_instructions_editor).pack(side="left")
    ttk.Button(button_row, text="閉じる(作業を始める)", command=win.destroy).pack(side="right")

    win.protocol("WM_DELETE_WINDOW", win.destroy)
    win.attributes("-topmost", True)
    win.lift()
    win.focus_force()
    root.wait_window(win)
    root.destroy()

def fix_kana(s):
    # 小書き「ヶ」を大書き「ケ」に統一する（例：雨ヶ谷→雨ケ谷）
    # ファイル名・店舗名の表記ゆれを吸収するための正規化
    return str(s).replace("ヶ", "ケ")

def half_upper(s):
    """先頭5文字比較用：全角英数字を半角に変換して大文字化"""
    return unicodedata.normalize('NFKC', str(s)).upper()

def norm(name):
    """店舗名を正規化（照合用）：前後空白・全半角スペース・「店」を除去、「ヶ」→「ケ」表記ゆれを統一"""
    return (fix_kana(name).strip()
            .replace("店", "")
            .replace(" ", "")
            .replace("　", ""))

def parse_bento_text_count(text):
    """弁当注文人数欄のテキスト表記（例：「放５児２」「児童２名」）に含まれる数値をすべて合計する。
    数値が1つも見つからない場合はNoneを返す。"""
    normalized = unicodedata.normalize('NFKC', str(text))
    numbers = re.findall(r'\d+', normalized)
    if not numbers:
        return None
    return sum(int(n) for n in numbers)

BENTO_COUNT_UNIT_KEYWORDS = ("児", "放", "名", "人")

def has_bento_count_unit(text):
    """「放５児２」「児童２名」のように人数を表す単位語(児/放/名/人)を含むかどうか。
    「①」のような記号や「あわせて」のような単位語のない文字列と、
    実際の人数表記を区別するために使う。"""
    return any(k in str(text) for k in BENTO_COUNT_UNIT_KEYWORDS)

def resolve_text_bento_count(text):
    """弁当注文人数欄の文字列を(人数, 記号扱いにするか)のペアにする。
    単位語を含み数値も抽出できる場合のみ実際の人数として合計し、
    それ以外（①等の記号、あわせて等の単位語なし文字列、単位語はあるが数値が
    見つからない場合）は人数として扱えないため0にし、記号扱いフラグを立てる
    （呼び出し側で出力セルを黄色にする）。"""
    if has_bento_count_unit(text):
        n = parse_bento_text_count(text)
        if n is not None:
            return n, False
    return 0, True

def to_file_link(path):
    """ローカルファイルパスをExcelハイパーリンク用のfile:///形式に変換する"""
    if not path:
        return ""
    return "file:///" + str(path).replace("\\", "/")

def lookup_shop(name, lookup_dict):
    """店舗名をマスタと照合して値を返す。
    「立替:」等の未知の接頭辞・接尾辞が付いていて完全一致しない場合は、
    マスタの店舗名が部分文字列として含まれる項目をフォールバックで採用する。"""
    key = norm(name)
    if key in lookup_dict:
        return lookup_dict[key]
    candidates = [k for k in lookup_dict if len(k) >= 3 and k in key]
    if candidates:
        return lookup_dict[max(candidates, key=len)]
    return ""

def load_url_map():
    url_map      = {}
    category_map = {}
    try:
        wb = openpyxl.load_workbook(REF_PATH, read_only=True, keep_vba=True, data_only=True)
        ws = wb["リスト"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            shop     = row[0]
            category = row[1]
            url      = row[3]
            if shop:
                url_map[str(shop).strip()]      = str(url).strip() if url else ""
                category_map[str(shop).strip()] = str(category).strip() if category else ""
        wb.close()
    except Exception as e:
        messagebox.showwarning("参照ファイル読み込みエラー", f"レク費urlリストを読み込めませんでした:\n{e}")
    return url_map, category_map

def parse_hug_text(text):
    """HUG「実費負担実績表」ページのテキストから店舗名・年月・日別の弁当【児】/【放】人数を読み取る。
    同じ日に複数件の記載がある場合はリストで保持する（例：{6: [1, 3]}）。"""
    lines = [ln.rstrip("\n") for ln in text.splitlines()]

    shop_raw = None
    for i, ln in enumerate(lines):
        if ln.strip() == "施設名":
            if i + 1 < len(lines):
                shop_raw = lines[i + 1].strip()
            break
    if not shop_raw:
        raise ValueError("施設名が見つかりませんでした。貼り付け内容を確認してください。")
    shop_display = shop_raw[:-1] if shop_raw.endswith("P") else shop_raw

    year = month = None
    for i, ln in enumerate(lines):
        if ln.strip() == "年月":
            try:
                year = int(lines[i + 1].strip())
                month = int(lines[i + 3].strip())
            except (IndexError, ValueError):
                pass
            break
    if not year or not month:
        raise ValueError("年月が見つかりませんでした。貼り付け内容を確認してください。")

    day_counts = {}
    current_day = None
    # 行末が1〜2桁の数字であれば日付区切りとして認識する（例：「海の日 20」のように
    # 祝日名が数字の前に付いている行も、行全体が数字のみの行と同様に日付として扱う）
    day_line_re = re.compile(r"^(?:\S*\s)?(\d{1,2})$")
    bento_re = re.compile(r"弁当【[児放]】.*?（(\d+)人）")
    for ln in lines:
        s = ln.strip()
        m_day = day_line_re.match(s)
        if m_day:
            d = int(m_day.group(1))
            if 1 <= d <= 31:
                current_day = d
            continue
        m = bento_re.search(s)
        if m and current_day is not None:
            count = int(m.group(1))
            if count > 0:
                day_counts.setdefault(current_day, []).append(count)

    return shop_display, year, month, day_counts


def get_shop_candidates(xlsx_path, shop_display, month):
    """対象月のデータ行からD列（店舗名）の一覧を取得し、shop_displayを部分一致で含む候補を返す。
    戻り値: (シート内の全店舗名リスト, 部分一致した候補リスト)"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    date_re = re.compile(r"^(\d{1,2})/(\d{1,2})$")
    names = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        shop_val = row[3] if len(row) > 3 else None
        date_val = row[4] if len(row) > 4 else None
        if not shop_val or not date_val:
            continue
        m = date_re.match(str(date_val).strip())
        if not m or int(m.group(1)) != month:
            continue
        name = str(shop_val).strip()
        if name not in seen:
            seen.add(name)
            names.append(name)
    wb.close()

    candidates = [n for n in names if shop_display in n]
    return names, candidates


def ask_shop_choice(parent, names, shop_display, ambiguous):
    """該当店舗が0件または複数件の場合に、手動で正しい店舗名を選ばせるダイアログ。
    選択されなければNoneを返す。"""
    result = {"value": None}
    win = tk.Toplevel(parent)
    win.title("店舗の選択")
    win.geometry("360x420")
    win.attributes("-topmost", True)
    win.grab_set()

    if ambiguous:
        msg = f"「{shop_display}」に該当する店舗候補が複数あります。正しい店舗名を選んでください。"
    else:
        msg = f"「{shop_display}」に該当する店舗が見つかりませんでした。正しい店舗名を選んでください。"
    tk.Label(win, text=msg, wraplength=340, justify="left").pack(padx=12, pady=(12, 6), anchor="w")

    search_var = tk.StringVar()
    search_row = ttk.Frame(win)
    search_row.pack(fill="x", padx=12, pady=(0, 6))
    ttk.Label(search_row, text="絞り込み:").pack(side="left")
    search_entry = ttk.Entry(search_row, textvariable=search_var)
    search_entry.pack(side="left", fill="x", expand=True, padx=(6, 0))

    list_frame = ttk.Frame(win)
    list_frame.pack(fill="both", expand=True, padx=12, pady=6)
    listbox = tk.Listbox(list_frame, exportselection=False)
    scroll = ttk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
    listbox.config(yscrollcommand=scroll.set)
    scroll.pack(side="right", fill="y")
    listbox.pack(side="left", fill="both", expand=True)

    def refresh_list(*_args):
        keyword = norm(search_var.get())
        listbox.delete(0, "end")
        for n in names:
            if not keyword or keyword in norm(n):
                listbox.insert("end", n)
        if listbox.size() > 0:
            listbox.selection_set(0)
            listbox.activate(0)

    def move_selection(delta):
        if listbox.size() == 0:
            return
        sel = listbox.curselection()
        idx = sel[0] + delta if sel else 0
        idx = max(0, min(listbox.size() - 1, idx))
        listbox.selection_clear(0, "end")
        listbox.selection_set(idx)
        listbox.activate(idx)
        listbox.see(idx)

    search_var.trace_add("write", refresh_list)
    refresh_list()

    search_entry.bind("<Up>", lambda e: (move_selection(-1), "break")[1])
    search_entry.bind("<Down>", lambda e: (move_selection(1), "break")[1])

    def on_ok():
        sel = listbox.curselection()
        if not sel:
            messagebox.showwarning("未選択", "店舗を選択してください。", parent=win)
            return
        result["value"] = listbox.get(sel[0])
        win.destroy()

    def on_cancel():
        result["value"] = None
        win.destroy()

    search_entry.bind("<Return>", lambda e: on_ok())
    listbox.bind("<Return>", lambda e: on_ok())
    listbox.bind("<Double-Button-1>", lambda e: on_ok())

    button_row = ttk.Frame(win)
    button_row.pack(side="bottom", fill="x", padx=12, pady=(0, 12))
    ttk.Button(button_row, text="OK", command=on_ok).pack(side="left")
    ttk.Button(button_row, text="キャンセル", command=on_cancel).pack(side="right")

    win.protocol("WM_DELETE_WINDOW", on_cancel)
    win.lift()
    win.focus_force()
    search_entry.focus_set()
    parent.wait_window(win)
    return result["value"]


def sync_hyperlink_refs(ws):
    """セルが保持するHyperlinkオブジェクトのref属性を、セルの実際の座標に合わせ直す。
    ws.insert_rows()はセルの位置(row/column)を移動させるが、セルに紐づくHyperlink
    オブジェクトのref（保存時にどのセルへリンクを紐付けるかを決める文字列）は追従して
    更新されない。ws._hyperlinksはload_workbook直後は空リストで、保存時にセルの
    hyperlink属性から再構築されるだけなので、そこを書き換えても意味が無い。
    各セルのhyperlink.refを直接cell.coordinateに合わせることで、保存後にA列リンクが
    1行ずれて別の店舗のファイルを開いてしまう不具合を防ぐ。"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink is not None:
                cell.hyperlink.ref = cell.coordinate

def apply_hug_data_to_excel(xlsx_path, shop_display, shop_full_name, year, month, day_counts):
    """G列に判定(〇/×)、H列にHUG内訳、I列に小計を書き込む。
    shop_full_nameはD列（店舗名）と完全一致させる対象店舗名（呼び出し側で確定済み）。
    H列の表示にはテキストから読み取った短い店舗名(shop_display)を使う。"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    ws.cell(row=1, column=7, value="判定")
    ws.cell(row=1, column=8, value="HUG")
    ws.cell(row=1, column=9, value="HUG小計")

    date_re = re.compile(r"^(\d{1,2})/(\d{1,2})$")
    matched_days = set()
    matched = mismatched = 0

    for row_idx in range(2, ws.max_row + 1):
        shop_val = ws.cell(row=row_idx, column=4).value
        date_val = ws.cell(row=row_idx, column=5).value
        count_val = ws.cell(row=row_idx, column=6).value
        if not shop_val or str(shop_val).strip() != shop_full_name:
            continue
        if not date_val:
            continue
        m = date_re.match(str(date_val).strip())
        if not m:
            continue
        row_month, row_day = int(m.group(1)), int(m.group(2))
        if row_month != month:
            continue

        counts_list = day_counts.get(row_day, [0])
        sum_counts  = sum(counts_list)
        h_text = f"{shop_display}{row_month}/{row_day}_" + "+".join(str(c) for c in counts_list) + "人"

        ws.cell(row=row_idx, column=8, value=h_text)
        ws.cell(row=row_idx, column=9, value=sum_counts)
        if count_val == sum_counts:
            ws.cell(row=row_idx, column=7, value="〇")
            matched += 1
        else:
            ws.cell(row=row_idx, column=7, value="×")
            mismatched += 1
        matched_days.add(row_day)

    inserted = 0
    unmatched_txt_days = sorted(set(day_counts) - matched_days)
    remaining_days = list(unmatched_txt_days)

    for day in remaining_days:
        block_rows = [r for r in range(2, ws.max_row + 1)
                      if ws.cell(row=r, column=4).value and str(ws.cell(row=r, column=4).value).strip() == shop_full_name]
        if not block_rows:
            continue  # この店舗の行が1つも無いため挿入位置を決められない

        insert_row = block_rows[-1] + 1
        for r in block_rows:
            date_val = ws.cell(row=r, column=5).value
            m = date_re.match(str(date_val).strip()) if date_val else None
            if m:
                r_month, r_day = int(m.group(1)), int(m.group(2))
                if (r_month, r_day) > (month, day):
                    insert_row = r
                    break

        ws.insert_rows(insert_row)
        ws.cell(row=insert_row, column=1, value=None)  # A列（ファイルリンク）は空欄のまま
        ws.cell(row=insert_row, column=4, value=shop_full_name)
        ws.cell(row=insert_row, column=5, value=f"{month}/{day}")

        counts_list = day_counts[day]
        sum_counts  = sum(counts_list)
        h_text = f"{shop_display}{month}/{day}_" + "+".join(str(c) for c in counts_list) + "人"
        ws.cell(row=insert_row, column=7, value="×")
        ws.cell(row=insert_row, column=8, value=h_text)
        ws.cell(row=insert_row, column=9, value=sum_counts)

        inserted += 1
        unmatched_txt_days.remove(day)

    sync_hyperlink_refs(ws)
    wb.save(xlsx_path)
    return matched, mismatched, inserted, unmatched_txt_days


def run_hug_paste_tool(root):
    win = tk.Toplevel(root)
    win.title("弁当代HUGテキスト転記")
    win.geometry("560x520")

    tk.Label(
        win,
        text="HUGの「実費負担実績表」ページのテキストを下に貼り付けて「転記」を押してください。",
        wraplength=520, justify="left"
    ).pack(padx=14, pady=(14, 6), anchor="w")

    PATH_PLACEHOLDER = "(未選択)"
    path_var = tk.StringVar(value=PATH_PLACEHOLDER)

    def choose_target_file():
        selected = filedialog.askopenfilename(
            title="転記先の集計表エクセルを選択してください",
            filetypes=[("Excel files", "*.xlsx")], parent=win
        )
        if selected:
            path_var.set(selected)

    def open_target_file():
        target = path_var.get()
        if not target or target == PATH_PLACEHOLDER:
            messagebox.showwarning("未選択", "転記先のエクセルが選択されていません。", parent=win)
            return
        try:
            os.startfile(target)
        except OSError as e:
            messagebox.showerror("開けませんでした", f"{target}\n{e}", parent=win)

    path_row = ttk.Frame(win)
    path_row.pack(fill="x", padx=14, pady=(0, 6))
    ttk.Label(path_row, text="転記先:").pack(side="left")
    ttk.Label(path_row, textvariable=path_var, foreground="#2f6f63", wraplength=380).pack(
        side="left", padx=(6, 6), fill="x", expand=True
    )
    ttk.Button(path_row, text="変更", command=choose_target_file).pack(side="right")
    ttk.Button(path_row, text="開く", command=open_target_file).pack(side="right", padx=(0, 6))

    choose_target_file()

    text_frame = ttk.Frame(win)
    text_frame.pack(fill="both", expand=True, padx=14, pady=6)
    text_widget = tk.Text(text_frame, wrap="word")
    scroll = ttk.Scrollbar(text_frame, orient="vertical", command=text_widget.yview)
    text_widget.config(yscrollcommand=scroll.set)
    scroll.pack(side="right", fill="y")
    text_widget.pack(side="left", fill="both", expand=True)

    context_menu = tk.Menu(text_widget, tearoff=0)
    context_menu.add_command(label="貼り付け", command=lambda: text_widget.event_generate("<<Paste>>"))
    context_menu.add_command(label="切り取り", command=lambda: text_widget.event_generate("<<Cut>>"))
    context_menu.add_command(label="コピー", command=lambda: text_widget.event_generate("<<Copy>>"))

    def show_context_menu(event):
        context_menu.tk_popup(event.x_root, event.y_root)

    text_widget.bind("<Button-3>", show_context_menu)

    def do_transfer():
        content = text_widget.get("1.0", "end")
        if not content.strip():
            messagebox.showwarning("入力なし", "テキストが貼り付けられていません。", parent=win)
            return
        try:
            shop_display, year, month, day_counts = parse_hug_text(content)
        except ValueError as e:
            messagebox.showerror("読み取りエラー", str(e), parent=win)
            return
        if not day_counts:
            messagebox.showwarning("弁当データなし", "テキストから弁当【児】のデータが見つかりませんでした。", parent=win)
            return

        xlsx_path = path_var.get()
        if not xlsx_path or xlsx_path == PATH_PLACEHOLDER:
            messagebox.showwarning("転記先未選択", "転記先のエクセルが選択されていません。「変更」ボタンから選択してください。", parent=win)
            return

        def offer_open_target_excel():
            if messagebox.askyesno("確認", "エクセルファイルを開きますか？", parent=win):
                try:
                    os.startfile(xlsx_path)
                except OSError as e:
                    messagebox.showerror("開けませんでした", f"{xlsx_path}\n{e}", parent=win)

        try:
            shop_names, candidates = get_shop_candidates(xlsx_path, shop_display, month)
        except Exception as e:
            messagebox.showerror("読み込みエラー", f"エクセルの店舗名取得に失敗しました:\n{e}", parent=win)
            offer_open_target_excel()
            return

        if len(candidates) == 1:
            shop_full_name = candidates[0]
        else:
            if not shop_names:
                messagebox.showerror("店舗なし", "エクセル内に対象月のデータが見つかりませんでした。", parent=win)
                offer_open_target_excel()
                return
            shop_full_name = ask_shop_choice(win, shop_names, shop_display, ambiguous=len(candidates) > 1)
            if not shop_full_name:
                return  # キャンセル時は転記しない（テキストは残す）

        try:
            matched, mismatched, inserted, unmatched_txt_days = apply_hug_data_to_excel(
                xlsx_path, shop_display, shop_full_name, year, month, day_counts
            )
        except Exception as e:
            messagebox.showerror("転記エラー", f"転記中にエラーが発生しました:\n{e}", parent=win)
            offer_open_target_excel()
            return

        msg = f"{shop_display}（{year}年{month}月）\n一致: {matched}件\n不一致: {mismatched}件"
        if inserted:
            msg += f"\n該当行なしのため新規挿入: {inserted}件"
        if unmatched_txt_days:
            msg += "\n店舗の行が見つからず挿入できなかった日付: " + ", ".join(f"{month}/{d}" for d in unmatched_txt_days)

        messagebox.showinfo("転記完了", msg, parent=win)
        text_widget.delete("1.0", "end")

        if matched == 0 and mismatched > 0:
            offer_open_target_excel()

    button_row = ttk.Frame(win)
    button_row.pack(side="bottom", fill="x", padx=14, pady=(0, 14))
    ttk.Button(button_row, text="転記", command=do_transfer).pack(side="left")
    ttk.Button(button_row, text="クリヤー", command=lambda: text_widget.delete("1.0", "end")).pack(side="left", padx=(8, 0))
    ttk.Button(button_row, text="閉じる", command=win.destroy).pack(side="right")

    win.protocol("WM_DELETE_WINDOW", win.destroy)
    win.lift()
    win.focus_force()
    root.wait_window(win)


def show_mode_launcher(root):
    """起動直後にどちらの作業を行うか選ばせる。"""
    result = {"mode": "exit"}
    win = tk.Toplevel(root)
    win.title("メニュー")
    win.geometry("360x210")
    win.attributes("-topmost", True)

    tk.Label(win, text="行う作業を選んでください", pady=10).pack()

    def choose(mode):
        result["mode"] = mode
        win.destroy()

    ttk.Button(win, text="① 弁当代を集計する", command=lambda: choose("aggregate")).pack(fill="x", padx=30, pady=6)
    ttk.Button(win, text="② 弁当代エクセルにHUGデータ連結", command=lambda: choose("hug")).pack(fill="x", padx=30, pady=6)
    ttk.Button(win, text="終了", command=lambda: choose("exit")).pack(fill="x", padx=30, pady=(14, 6))

    win.protocol("WM_DELETE_WINDOW", lambda: choose("exit"))
    win.lift()
    win.focus_force()
    root.wait_window(win)
    return result["mode"]


def save_shop_list_excel(final_df, zero_shops, missing_shops, save_path):
    """集計表(final_df/zero_shops/missing_shops)に登場する店舗をまとめ、
    同じ店舗が複数行(日付ごと等)に渡って重複していても1行だけになる店舗一覧を
    別ファイルとして保存する。列は「ファイルリンク・区分・店舗名」の3列。"""
    parts = []
    if not final_df.empty:
        tmp = final_df[['【注文あり店舗】ファイルリンク', '区分', '店舗名']].copy()
        tmp.columns = ['ファイルリンク', '区分', '店舗名']
        parts.append(tmp)
    for shop in zero_shops:
        parts.append(pd.DataFrame([{
            'ファイルリンク': shop.get('url', ''),
            '区分':       shop.get('区分', ''),
            '店舗名':     shop.get('店舗名', ''),
        }]))
    for shop in missing_shops:
        parts.append(pd.DataFrame([{
            'ファイルリンク': shop.get('url', ''),
            '区分':       shop.get('区分', ''),
            '店舗名':     shop.get('店舗名', ''),
        }]))

    if not parts:
        return

    shop_df = pd.concat(parts, ignore_index=True)
    shop_df = shop_df.drop_duplicates(subset=['ファイルリンク', '区分', '店舗名'], keep='first').reset_index(drop=True)

    shop_df.to_excel(save_path, index=False)

    wb = openpyxl.load_workbook(save_path)
    ws = wb.active

    hyperlink_font = Font(color="0563C1", underline="single")
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        url_val = cell.value
        if url_val and str(url_val).startswith(("http", "file:")):
            cell.hyperlink = url_val
            cell.value     = url_val
            cell.font      = hyperlink_font

    ws.column_dimensions['A'].width = 4
    for col_letter, col_idx in [('B', 2), ('C', 3)]:
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value))
             for r in range(2, ws.max_row + 1)
             if ws.cell(row=r, column=col_idx).value not in (None, "")),
            default=8
        )
        ws.column_dimensions[col_letter].width = max_len + 2

    for cell in ws[1]:
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    wb.save(save_path)


def main():
    minimize_console_window()
    show_instructions_dialog()

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    while True:
        mode = show_mode_launcher(root)
        if mode == "exit":
            break
        if mode == "hug":
            run_hug_paste_tool(root)
            continue

        url_map, category_map = load_url_map()
        url_lookup      = {norm(k): v for k, v in url_map.items()}       # 照合用（「店」除去済みキー→url）
        category_lookup = {norm(k): v for k, v in category_map.items()}  # 照合用（「店」除去済みキー→区分）

        now = datetime.now()
        if now.month == 1:
            default_ym = f"{now.year - 1}12"
        else:
            default_ym = f"{now.year}{now.month - 1:02d}"

        target_ym = simpledialog.askstring(
            "入力", "抽出する年月を入力してください\n（例: 202604）",
            initialvalue=default_ym, parent=root
        )
        if not target_ym: continue

        folder_path = filedialog.askdirectory(title="弁当代エクセルのフォルダを選択してください")
        if not folder_path: continue

        target_files = [f for f in os.listdir(folder_path)
                        if (f.endswith(".xlsx") or f.endswith(".xls")) and not f.startswith("~$")]

        combined_list = []
        zero_shops    = []
        found_shop_names = set()  # 処理済み店舗名を記録

        progress_window = tk.Toplevel(root)
        progress_window.title("データ集計中...")
        progress_window.geometry("450x150")
        progress_window.attributes("-topmost", True)
        label = tk.Label(progress_window, text="準備中...", wraplength=400)
        label.pack(pady=10)
        pb = ttk.Progressbar(progress_window, orient="horizontal", length=350, mode="determinate")
        pb.pack(pady=5)
        pb["maximum"] = len(target_files)

        for i, filename in enumerate(target_files):
            pb["value"] = i + 1
            label.config(text=f"処理中 ({i+1}/{len(target_files)}):\n{filename}")
            progress_window.update()

            file_path = os.path.join(folder_path, filename)
            try:
                xl = pd.ExcelFile(file_path, engine='openpyxl')
                if target_ym not in xl.sheet_names: continue

                wb_in = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                ws_in = wb_in[target_ym]
                shop_name = ws_in["B3"].value if ws_in["B3"].value else None
                if shop_name:
                    shop_name = str(shop_name).replace("メソッド", "").replace("パーク", "")
                if not shop_name:
                    import re
                    m = re.search(r'【(.+?)】', filename)
                    shop_name = m.group(1) if m else "名称不明"
                wb_in.close()

                # 店舗名の「ヶ」→「ケ」表記ゆれを統一する
                shop_name = fix_kana(shop_name)
                # ファイル名（出力表示用）の「ヶ」→「ケ」表記ゆれを統一する
                disp_filename = fix_kana(filename)

                found_shop_names.add(norm(shop_name))        # norm済みで登録
                found_shop_names.add(str(shop_name).strip()) # 元の名前でも登録

                df_raw = pd.read_excel(file_path, sheet_name=target_ym, header=None, engine='openpyxl')

                target_row_idx, target_col_idx, date_col_idx = None, None, None
                for r_idx, row in df_raw.iterrows():
                    row_str_list = [str(x).replace(" ", "").replace("　", "") for x in row]
                    found_col = next((c_idx for c_idx, val in enumerate(row_str_list) if "弁当注文人数" in val), None)
                    if found_col is not None:
                        target_row_idx, target_col_idx = r_idx, found_col
                        date_col_idx = next((c_idx for c_idx, val in enumerate(row_str_list) if "日付" in val), None)
                        break

                if target_row_idx is None: continue

                df = pd.read_excel(file_path, sheet_name=target_ym, header=target_row_idx, engine='openpyxl')
                col_name_bento = df.columns[target_col_idx]
                col_name_date  = df.columns[date_col_idx] if date_col_idx is not None else "日付"

                def is_text_entry(x):
                    if pd.isna(x): return False
                    if isinstance(x, (int, float)): return False
                    s = str(x).strip()
                    if s == "": return False
                    try:
                        float(s)
                        return False
                    except ValueError:
                        return True

                mask_text  = df[col_name_bento].apply(is_text_entry)
                df_text    = df[mask_text].copy()

                df['_bento_num'] = pd.to_numeric(df[col_name_bento], errors='coerce')
                total        = df['_bento_num'].sum()
                mask_numeric = df['_bento_num'] > 0
                df_numeric   = df[mask_numeric].copy()
                df_numeric['_is_symbol'] = False

                resolved = df_text[col_name_bento].apply(resolve_text_bento_count)
                df_text['_bento_num'] = resolved.apply(lambda t: t[0])
                df_text['_is_symbol'] = resolved.apply(lambda t: t[1])
                df_text[col_name_bento] = df_text['_bento_num']
                df_numeric[col_name_bento] = df_numeric['_bento_num']

                df_filtered = pd.concat([df_numeric, df_text]).drop_duplicates()

                if df_filtered.empty and total <= 0:
                    zero_shops.append({
                        'ファイル名': disp_filename,  # 表記ゆれ修正済みのファイル名を使用
                        '店舗名': shop_name,           # 表記ゆれ修正済みの店舗名を使用
                        '区分': lookup_shop(shop_name, category_lookup),
                        'url': to_file_link(file_path)
                    })
                    continue

                if not df_filtered.empty:
                    df_filtered = df_filtered.copy()
                    df_filtered['ファイル名'] = disp_filename  # 表記ゆれ修正済みのファイル名を使用
                    df_filtered['店舗名']    = shop_name       # 表記ゆれ修正済みの店舗名を使用
                    df_filtered['区分']      = lookup_shop(shop_name, category_lookup)
                    df_filtered['ファイルリンク'] = to_file_link(file_path)
                    df_filtered['sort_date'] = pd.to_datetime(df_filtered[col_name_date], errors='coerce')

                    res_df = df_filtered[['ファイル名', '区分', '店舗名', col_name_date, col_name_bento, 'ファイルリンク', 'sort_date', '_is_symbol']]
                    res_df.columns = ['ファイル名', '区分', '店舗名', '日付', '弁当注文人数', 'ファイルリンク', 'sort_date', '_is_symbol']

                    # 同じ日付が複数行ある場合は1行にまとめ、弁当注文人数は合算する
                    def _sum_bento(s):
                        nums = pd.to_numeric(s, errors='coerce')
                        if nums.notna().all():
                            total = nums.sum()
                            return int(total) if float(total).is_integer() else total
                        return "+".join(str(v) for v in s)

                    res_df = res_df.groupby('日付', as_index=False, sort=False).agg({
                        'ファイル名':     'first',
                        '区分':         'first',
                        '店舗名':       'first',
                        '弁当注文人数':  _sum_bento,
                        'ファイルリンク': 'first',
                        'sort_date':    'first',
                        '_is_symbol':   'any',
                    })
                    res_df = res_df[['ファイル名', '区分', '店舗名', '日付', '弁当注文人数', 'ファイルリンク', 'sort_date', '_is_symbol']]

                    combined_list.append(res_df)
                else:
                    zero_shops.append({
                        'ファイル名': filename,
                        '店舗名': shop_name,
                        '区分': lookup_shop(shop_name, category_lookup),
                        'url': to_file_link(file_path)
                    })

            except:
                continue

        progress_window.destroy()

        # リストシートにあるがファイルが存在しない店舗
        missing_shops = [
            {'ファイル名': '', '店舗名': name, '区分': category_lookup.get(norm(name), ""), 'url': url_lookup.get(norm(name), "")}
            for name in url_map
            if norm(name) not in found_shop_names
            and name not in found_shop_names
        ]

        if combined_list or zero_shops or missing_shops:
            folder_name = os.path.basename(os.path.normpath(folder_path)).replace("　", "").replace(" ", "")
            output_name = f"{folder_name}集計表.xlsx"
            save_path = os.path.join(os.path.expanduser("~"), "Downloads", output_name)

            if combined_list:
                final_df = pd.concat(combined_list, ignore_index=True)
                final_df = final_df.sort_values(by=['区分', '店舗名', 'sort_date']).reset_index(drop=True)

                def format_md(x):
                    try:
                        if pd.isna(x): return ""
                        return f"{x.month}/{x.day}"
                    except:
                        return str(x)

                final_df['日付'] = final_df['sort_date'].apply(format_md)
                final_df = final_df.drop(columns=['sort_date'])

                # 「①」等の記号扱いで0にした行を、後でExcel保存後に黄色くするための一覧
                bento_symbol_flags = final_df['_is_symbol'].tolist()
                final_df = final_df.drop(columns=['_is_symbol'])

                final_df.insert(0, '【注文あり店舗】ファイルリンク', final_df['ファイルリンク'])
                final_df = final_df.drop(columns=['ファイルリンク'])
            else:
                final_df = pd.DataFrame(columns=['【注文あり店舗】ファイルリンク', 'ファイル名', '区分', '店舗名', '日付', '弁当注文人数'])
                bento_symbol_flags = []

            final_df.to_excel(save_path, index=False)

            wb = openpyxl.load_workbook(save_path)
            ws = wb.active

            # A列をハイパーリンクに変換
            hyperlink_font = Font(color="0563C1", underline="single")
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=1)
                url_val = cell.value
                if url_val and str(url_val).startswith(("http", "file:")):
                    cell.hyperlink = url_val
                    cell.value     = url_val
                    cell.font      = hyperlink_font

            # 縞模様（店舗名が同じ間は同じ色にする）
            fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            current_fill = False
            last_shop = None
            for row_idx in range(2, ws.max_row + 1):
                shop_val = str(ws.cell(row=row_idx, column=4).value)
                if shop_val != last_shop:
                    current_fill = not current_fill
                    last_shop = shop_val
                if current_fill:
                    for col_idx in range(1, 7):
                        ws.cell(row=row_idx, column=col_idx).fill = fill_gray

            # 「①」等の記号扱いで人数を0にした行は、弁当注文人数セル(F列)を黄色にして目立たせる
            fill_bento_symbol = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for offset, is_symbol in enumerate(bento_symbol_flags):
                if is_symbol:
                    ws.cell(row=2 + offset, column=6).fill = fill_bento_symbol

            ws.column_dimensions['A'].width  = 4
            ws.column_dimensions['B'].hidden = True
            ws.column_dimensions['F'].width  = 4

            # 弁当代HUGテキスト転記機能で使う列のヘッダーを先に用意しておく
            ws.cell(row=1, column=7, value="判定")
            ws.cell(row=1, column=8, value="HUG")
            ws.cell(row=1, column=9, value="HUG小計")

            # 末尾セクション共通処理
            def append_section_with_url(ws, shops, header_texts, header_color, row_color, url_key='url', name_key='店舗名', category_key='区分'):
                blank_row = ws.max_row + 2
                fill_h = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                fill_r = PatternFill(start_color=row_color,    end_color=row_color,    fill_type="solid")
                bold   = Font(bold=True)
                for col_idx, text in enumerate(header_texts, start=1):
                    cell = ws.cell(row=blank_row, column=col_idx, value=text)
                    cell.fill = fill_h
                    cell.font = bold
                for shop in shops:
                    blank_row += 1
                    url_val  = shop.get(url_key, "")
                    url_cell = ws.cell(row=blank_row, column=1)
                    url_cell.fill = fill_r
                    if url_val and str(url_val).startswith(("http", "file:")):
                        url_cell.hyperlink = url_val
                        url_cell.value     = url_val
                        url_cell.font      = Font(color="0563C1", underline="single")
                    col = 2
                    if 'ファイル名' in shop:
                        ws.cell(row=blank_row, column=col, value=shop['ファイル名']).fill = fill_r
                        col += 1
                    if category_key in shop:
                        ws.cell(row=blank_row, column=col, value=shop[category_key]).fill = fill_r
                        col += 1
                    ws.cell(row=blank_row, column=col, value=shop[name_key]).fill = fill_r

            if zero_shops:
                zero_shops = sorted(zero_shops, key=lambda s: (s.get('区分', ''), s.get('店舗名', '')))
                append_section_with_url(
                    ws, zero_shops,
                    ["【注文なし店舗】ファイルリンク", "ファイル名", "区分", "店舗名"],
                    "FFFF00", "FFE0E0"
                )
    #【ファイルなし店舗】レク費urlリストにあるがファイルが存在しない店舗のうち、店舗名が類似しているものを除外（例：先頭5文字で照合）
            if missing_shops:
                upper_prefixes = {
                    half_upper(str(n).strip()[:5])
                    for n in (set(final_df['店舗名']) | {s['店舗名'] for s in zero_shops})
                    if len(str(n).strip()) >= 5
                }
                missing_shops = [
                    s for s in missing_shops
                    if half_upper(str(s['店舗名']).strip()[:5]) not in upper_prefixes
                ]

            if missing_shops:
                missing_shops = sorted(missing_shops, key=lambda s: (s.get('区分', ''), s.get('店舗名', '')))
                append_section_with_url(
                    ws, missing_shops,
                    ["【ファイルなし店舗】レク費url", "ファイル名", "区分", "店舗名"],
                    "FFC000", "FFE8CC"
                )

            # C,D,E列を内容に合わせて幅自動調整
            for col_letter, col_idx in [('C', 3), ('D', 4), ('E', 5)]:
                max_len = max(
                    (len(str(ws.cell(row=r, column=col_idx).value))
                     for r in range(2, ws.max_row + 1)
                     if ws.cell(row=r, column=col_idx).value not in (None, "")),
                    default=8
                )
                ws.column_dimensions[col_letter].width = max_len + 2

            # ヘッダー行：折り返して全体を表示し、高さは自動調整のため明示指定しない
            for cell in ws[1]:
                cell.alignment = Alignment(wrap_text=True, vertical='center')

            wb.save(save_path)

            shop_list_name = "弁当代チェック_店舗一覧.xlsx"
            shop_list_path = os.path.join(os.path.expanduser("~"), "Downloads", shop_list_name)
            save_shop_list_excel(final_df, zero_shops, missing_shops, shop_list_path)

            messagebox.showinfo("完了", f"保存しました：\n{output_name}\n{shop_list_name}")
            os.startfile(shop_list_path)
        else:
            messagebox.showwarning("結果", "データが見つかりませんでした。")

    root.destroy()
    sys.exit()

if __name__ == "__main__":
    main()
