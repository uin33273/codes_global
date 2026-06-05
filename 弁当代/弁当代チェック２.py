#最初にxlsxファイルを取り込み、弁当注文人数の列を探して、0より大きい行だけ抽出して結合するコード
#"C:\Users\owner\Desktop\works\保管書類\codes\global_shoplist\table_HHHL共通店舗一覧m.xlsm"をマスタとして参照している
#出力は「弁当チェック_年月_yyyymmdd_hhmm.xlsx」という名前で、ダウンロードフォルダに保存される
#出力ファイルのA列には、マスタにURLがあればハイパーリンクで表示される
#店舗名がマスタにあるがファイルが存在しない店舗は、末尾のセクションに「【ファイルなし店舗】レク費urlリストにあるがファイルが存在しない店舗」としてまとめて表示される（ただし、店舗名が類似しているものは除外される）
#店舗名が空欄のファイルは、ファイル名から店舗名を推測して照合する（例：ファイル名に「【○○店】」のように店舗名が含まれていれば、それを店舗名として扱う）
import pandas as pd
import openpyxl
import os
import zipfile
import tempfile
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox, ttk
import sys
import unicodedata
import warnings
from datetime import datetime
from openpyxl.styles import PatternFill, Font

warnings.simplefilter('ignore', UserWarning)

REF_PATH = r"C:\Users\owner\Desktop\works\保管書類\codes\global_shoplist\table_HHHL共通店舗一覧m.xlsm"

def fix_kana(s):
    # 小書き「ヶ」を大書き「ケ」に統一する（例：雨ヶ谷→雨ケ谷）
    # ファイル名・店舗名の表記ゆれを吸収するための正規化
    return str(s).replace("ヶ", "ケ")

def half_upper(s):
    """先頭5文字比較用：全角英数字を半角に変換して大文字化"""
    return unicodedata.normalize('NFKC', str(s)).upper()

def norm(name):
    """店舗名を正規化（照合用）：前後空白・全半角スペース・「店」を除去"""
    return (str(name).strip()
            .replace("店", "")
            .replace(" ", "")
            .replace("　", ""))

def load_url_map():
    url_map = {}
    try:
        wb = openpyxl.load_workbook(REF_PATH, read_only=True, keep_vba=True, data_only=True)
        ws = wb["リスト"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            shop = row[0]
            url  = row[3]
            if shop:
                url_map[str(shop).strip()] = str(url).strip() if url else ""
        wb.close()
    except Exception as e:
        messagebox.showwarning("参照ファイル読み込みエラー", f"レク費urlリストを読み込めませんでした:\n{e}")
    return url_map

def main():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    url_map    = load_url_map()
    url_lookup = {norm(k): v for k, v in url_map.items()}  # 照合用（「店」除去済みキー→url）

    target_ym = simpledialog.askstring("入力", "抽出する年月を入力してください\n（例: 202604）", parent=root)
    if not target_ym: return

    # フォルダまたはZIP選択ダイアログ
    choice_win = tk.Toplevel(root)
    choice_win.title("入力元選択")
    choice_win.geometry("400x110")
    choice_win.attributes("-topmost", True)
    choice_win.resizable(False, False)
    choice_win.grab_set()

    folder_path = [None]
    tmpdir      = [None]

    def pick_folder():
        path = filedialog.askdirectory(title="エクセルがあるフォルダを選択してください")
        if path:
            folder_path[0] = path
        choice_win.destroy()

    def pick_zip():
        path = filedialog.askopenfilename(
            title="エクセルがあるZIPファイルを選択してください",
            filetypes=[("ZIPファイル", "*.zip")]
        )
        if path:
            td = tempfile.mkdtemp()
            with zipfile.ZipFile(path, 'r') as zf:
                zf.extractall(td)
            tmpdir[0] = td
            found_dir = td
            for dirpath, _, filenames in os.walk(td):
                if any(f.endswith((".xlsx", ".xls")) and not f.startswith("~$") for f in filenames):
                    found_dir = dirpath
                    break
            folder_path[0] = found_dir
        choice_win.destroy()

    tk.Label(choice_win, text="エクセルがあるフォルダまたはZIPファイルを選択してください", pady=8).pack()
    btn_frame = tk.Frame(choice_win)
    btn_frame.pack()
    tk.Button(btn_frame, text="フォルダを選択",    width=18, height=2, command=pick_folder).pack(side=tk.LEFT, padx=12)
    tk.Button(btn_frame, text="ZIPファイルを選択", width=18, height=2, command=pick_zip).pack(side=tk.LEFT, padx=12)

    choice_win.wait_window()
    if not folder_path[0]: return
    folder_path = folder_path[0]

    try:
        target_files = [f for f in os.listdir(folder_path)
                        if (f.endswith(".xlsx") or f.endswith(".xls")) and not f.startswith("~$")]

        combined_list = []
        zero_shops    = []
        found_shop_names = set()  # 処理済み店舗名を記録

        progress_window = tk.Toplevel(root)
        progress_window.title("データ集計中...")
        progress_window.geometry("450x150")
        progress_window.attributes("-topmost", True)
        label = tk.Label(progress_window, text="準備中...", wraplength=400)
        label.pack(pady=10)
        pb = ttk.Progressbar(progress_window, orient="horizontal", length=350, mode="determinate")
        pb.pack(pady=5)
        pb["maximum"] = len(target_files)

        for i, filename in enumerate(target_files):
            pb["value"] = i + 1
            label.config(text=f"処理中 ({i+1}/{len(target_files)}):\n{filename}")
            progress_window.update()

            file_path = os.path.join(folder_path, filename)
            try:
                xl = pd.ExcelFile(file_path, engine='openpyxl')
                if target_ym not in xl.sheet_names: continue

                wb_in = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                ws_in = wb_in[target_ym]
                shop_name = ws_in["B3"].value if ws_in["B3"].value else None
                if not shop_name:
                    import re
                    m = re.search(r'【(.+?)】', filename)
                    shop_name = m.group(1) if m else "名称不明"
                wb_in.close()

                # 店舗名の「ヶ」→「ケ」表記ゆれを統一する
                shop_name = fix_kana(shop_name)
                # ファイル名（出力表示用）の「ヶ」→「ケ」表記ゆれを統一する
                disp_filename = fix_kana(filename)

                found_shop_names.add(norm(shop_name))        # norm済みで登録
                found_shop_names.add(str(shop_name).strip()) # 元の名前でも登録

                df_raw = pd.read_excel(file_path, sheet_name=target_ym, header=None, engine='openpyxl')

                target_row_idx, target_col_idx, date_col_idx = None, None, None
                for r_idx, row in df_raw.iterrows():
                    row_str_list = [str(x).replace(" ", "").replace("　", "") for x in row]
                    found_col = next((c_idx for c_idx, val in enumerate(row_str_list) if "弁当注文人数" in val), None)
                    if found_col is not None:
                        target_row_idx, target_col_idx = r_idx, found_col
                        date_col_idx = next((c_idx for c_idx, val in enumerate(row_str_list) if "日付" in val), None)
                        break

                if target_row_idx is None: continue

                df = pd.read_excel(file_path, sheet_name=target_ym, header=target_row_idx, engine='openpyxl')
                col_name_bento = df.columns[target_col_idx]
                col_name_date  = df.columns[date_col_idx] if date_col_idx is not None else "日付"

                def is_text_entry(x):
                    if pd.isna(x): return False
                    if isinstance(x, (int, float)): return False
                    s = str(x).strip()
                    if s == "": return False
                    try:
                        float(s)
                        return False
                    except ValueError:
                        return True

                mask_text  = df[col_name_bento].apply(is_text_entry)
                df_text    = df[mask_text].copy()

                df['_bento_num'] = pd.to_numeric(df[col_name_bento], errors='coerce')
                total        = df['_bento_num'].sum()
                mask_numeric = df['_bento_num'] > 0
                df_numeric   = df[mask_numeric].copy()

                df_text['_bento_num']      = df_text[col_name_bento]
                df_numeric[col_name_bento] = df_numeric['_bento_num']

                df_filtered = pd.concat([df_numeric, df_text]).drop_duplicates()

                if df_filtered.empty and total <= 0:
                    zero_shops.append({
                        'ファイル名': disp_filename,  # 表記ゆれ修正済みのファイル名を使用
                        '店舗名': shop_name,           # 表記ゆれ修正済みの店舗名を使用
                        'url': url_lookup.get(norm(shop_name), "")
                    })
                    continue

                if not df_filtered.empty:
                    df_filtered = df_filtered.copy()
                    df_filtered['ファイル名'] = disp_filename  # 表記ゆれ修正済みのファイル名を使用
                    df_filtered['店舗名']    = shop_name       # 表記ゆれ修正済みの店舗名を使用
                    df_filtered['sort_date'] = pd.to_datetime(df_filtered[col_name_date], errors='coerce')

                    res_df = df_filtered[['ファイル名', '店舗名', col_name_date, col_name_bento, 'sort_date']]
                    res_df.columns = ['ファイル名', '店舗名', '日付', '弁当注文人数', 'sort_date']
                    combined_list.append(res_df)
                else:
                    zero_shops.append({
                        'ファイル名': filename,
                        '店舗名': shop_name,
                        'url': url_lookup.get(norm(shop_name), "")
                    })

            except:
                continue

        progress_window.destroy()

        # リストシートにあるがファイルが存在しない店舗
        missing_shops = [
            {'ファイル名': '', '店舗名': name, 'url': url_lookup.get(norm(name), "")}
            for name in url_map
            if norm(name) not in found_shop_names
            and name not in found_shop_names
        ]

        if combined_list or zero_shops or missing_shops:
            now = datetime.now().strftime("%Y%m%d_%H%M")
            output_name = f"弁当チェック_{target_ym}_{now}.xlsx"
            save_path = os.path.join(os.path.expanduser("~"), "Downloads", output_name)

            if combined_list:
                final_df = pd.concat(combined_list, ignore_index=True)
                final_df = final_df.sort_values(by=['ファイル名', 'sort_date']).reset_index(drop=True)

                def format_md(x):
                    try:
                        if pd.isna(x): return ""
                        return f"{x.month}/{x.day}"
                    except:
                        return str(x)

                final_df['日付'] = final_df['sort_date'].apply(format_md)
                final_df = final_df.drop(columns=['sort_date'])

                final_df.insert(0, '【注文あり店舗】レク費url', final_df['店舗名'].apply(
                    lambda s: url_lookup.get(norm(s), "")
                ))
            else:
                final_df = pd.DataFrame(columns=['【注文あり店舗】レク費url', 'ファイル名', '店舗名', '日付', '弁当注文人数'])

            final_df.to_excel(save_path, index=False)

            wb = openpyxl.load_workbook(save_path)
            ws = wb.active

            # A列をハイパーリンクに変換
            hyperlink_font = Font(color="0563C1", underline="single")
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=1)
                url_val = cell.value
                if url_val and str(url_val).startswith("http"):
                    cell.hyperlink = url_val
                    cell.value     = url_val
                    cell.font      = hyperlink_font

            # 縞模様
            fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            current_fill = False
            last_date = None
            for row_idx in range(2, ws.max_row + 1):
                date_val = str(ws.cell(row=row_idx, column=4).value)
                if date_val != last_date:
                    current_fill = not current_fill
                    last_date = date_val
                if current_fill:
                    for col_idx in range(1, 6):
                        ws.cell(row=row_idx, column=col_idx).fill = fill_gray

            ws.column_dimensions['A'].width  = 10
            ws.column_dimensions['B'].hidden = True
            ws.column_dimensions['C'].width  = 30
            ws.column_dimensions['D'].width  = 10
            ws.column_dimensions['E'].width  = 15

            # 末尾セクション共通処理
            def append_section_with_url(ws, shops, header_texts, header_color, row_color, url_key='url', name_key='店舗名'):
                blank_row = ws.max_row + 2
                fill_h = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                fill_r = PatternFill(start_color=row_color,    end_color=row_color,    fill_type="solid")
                bold   = Font(bold=True)
                for col_idx, text in enumerate(header_texts, start=1):
                    cell = ws.cell(row=blank_row, column=col_idx, value=text)
                    cell.fill = fill_h
                    cell.font = bold
                for shop in shops:
                    blank_row += 1
                    url_val  = shop.get(url_key, "")
                    url_cell = ws.cell(row=blank_row, column=1)
                    url_cell.fill = fill_r
                    if url_val and str(url_val).startswith("http"):
                        url_cell.hyperlink = url_val
                        url_cell.value     = url_val
                        url_cell.font      = Font(color="0563C1", underline="single")
                    col = 2
                    if 'ファイル名' in shop:
                        ws.cell(row=blank_row, column=col, value=shop['ファイル名']).fill = fill_r
                        col += 1
                    ws.cell(row=blank_row, column=col, value=shop[name_key]).fill = fill_r

            if zero_shops:
                append_section_with_url(
                    ws, zero_shops,
                    ["【注文なし店舗】レク費url", "ファイル名", "店舗名"],
                    "FFFF00", "FFE0E0"
                )
#【ファイルなし店舗】レク費urlリストにあるがファイルが存在しない店舗のうち、店舗名が類似しているものを除外（例：先頭5文字で照合）
            if missing_shops:
                upper_prefixes = {
                    half_upper(str(n).strip()[:5])
                    for n in (set(final_df['店舗名']) | {s['店舗名'] for s in zero_shops})
                    if len(str(n).strip()) >= 5
                }
                missing_shops = [
                    s for s in missing_shops
                    if half_upper(str(s['店舗名']).strip()[:5]) not in upper_prefixes
                ]

            if missing_shops:
                append_section_with_url(
                    ws, missing_shops,
                    ["【ファイルなし店舗】レク費url", "ファイル名", "店舗名"],
                    "FFC000", "FFE8CC"
                )

            wb.save(save_path)
            messagebox.showinfo("完了", f"保存しました：\n{output_name}")
            os.startfile(save_path)
        else:
            messagebox.showwarning("結果", "データが見つかりませんでした。")

    finally:
        if tmpdir[0] and os.path.exists(tmpdir[0]):
            import shutil
            shutil.rmtree(tmpdir[0], ignore_errors=True)

    root.destroy()
    sys.exit()

if __name__ == "__main__":
    main()