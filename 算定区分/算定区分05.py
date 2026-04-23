#このコードは、2つのExcelファイルを読み込み、特定の列を基に新しいシートに数式を入力して保存するものです。
#1つ目のファイルは「算定区分・加算等集計表（運営用）」で、2つ目のファイルは「2-スプレッドシートの店名へ紐づけ」です。
#最終的に、1つ目のファイルの店舗名を基に、2つ目のファイルの店舗名列と照合し、数式を入力した新しいシート「LOOKUP」を作成して保存します。
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

# --- 共通関数：見出し名から列番号を探す ---
def find_column_by_header(ws, header_row, target_name):
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val == target_name:
            return c
    return None

def force_quit(event=None):
    """Escキーで呼び出される強制終了処理"""
    messagebox.showwarning("強制終了", "プログラムを強制終了します。")
    os._exit(0)

class FinalProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("算定区分05: Excelコピペ用データ作成")
        self.root.geometry("300x120")
        self.root.attributes('-topmost', True)
        self.root.bind("<Escape>", lambda e: self.root.destroy())
        
        self.label = tk.Label(root, text="Excel処理を実行中...\n(Escキーで強制終了)", pady=20)
        self.label.pack()

    def start_process(self):
        downloads = Path.home() / "Downloads"
        
        # --- 1. 最初のファイル（運営用集計表）の読み込み ---
        path_agg = filedialog.askopenfilename(
            title="1-算定区分・加算等集計表（運営用）を選択してください",
            initialdir=str(downloads),
            parent=self.root
        )
        if not path_agg:
            self.root.destroy()
            return

        # --- 2. 2番目のファイル（紐づけ済みファイル）の読み込み ---
        path_target = filedialog.askopenfilename(
            title="2-スプレッドシートの店名へ紐づけ を選択してください",
            initialdir=str(downloads),
            initialfile="2-スプレッドシートの店名へ紐づけ.xlsx",
            parent=self.root
        )
        if not path_target:
            self.root.destroy()
            return

        try:
            # データ処理開始
            wb_agg = openpyxl.load_workbook(path_agg, data_only=True)
            ws_agg = wb_agg["集計"]
            col_shop_agg = find_column_by_header(ws_agg, 3, "店舗名")
            
            if not col_shop_agg:
                messagebox.showerror("エラー", "集計表に『店舗名』が見つかりません。", parent=self.root)
                self.root.destroy()
                return

            shop_names = [ws_agg.cell(row=r, column=col_shop_agg).value 
                          for r in range(4, ws_agg.max_row + 1) 
                          if ws_agg.cell(row=r, column=col_shop_agg).value]
            wb_agg.close()

            wb_target = openpyxl.load_workbook(path_target)
            ws_mismatch = wb_target.active 
            ws_mismatch.title = "店舗名紐づけ"

            col_key_target = find_column_by_header(ws_mismatch, 1, "店舗名")
            if not col_key_target:
                messagebox.showerror("エラー", "紐づけファイルに『店舗名』列がありません。", parent=self.root)
                self.root.destroy()
                return
            
            key_col_letter = get_column_letter(col_key_target)

            # LOOKUPシート作成
            if "LOOKUP" in wb_target.sheetnames: del wb_target["LOOKUP"]
            ws_lookup = wb_target.create_sheet("LOOKUP")
            ws_lookup.cell(row=1, column=1).value = "店舗名"
            for i, name in enumerate(shop_names, start=2):
                ws_lookup.cell(row=i, column=1).value = name

            # XLOOKUP数式の入力
            for c in range(1, 11): 
                header_val = ws_mismatch.cell(row=1, column=c).value
                ws_lookup.cell(row=1, column=c+1).value = header_val 
                t_col = get_column_letter(c) 
                for r in range(2, len(shop_names) + 2):
                    ws_lookup.cell(row=r, column=c+1).value = f'=_xlfn.XLOOKUP($A{r}, \'店舗名紐づけ\'!${key_col_letter}$2:${key_col_letter}$1000, \'店舗名紐づけ\'!${t_col}$2:${t_col}$1000, "")'

            # 保存
            save_path = os.path.join(str(downloads), "3-spreadsheetへコピペ用データ.xlsx")
            wb_target.active = ws_lookup
            wb_target.save(save_path)
            
            messagebox.showinfo("完了", f"すべての処理が完了しました！\n\n保存先:\n{save_path}", parent=self.root)

        except Exception as e:
            messagebox.showerror("エラー", f"予期せぬエラー:\n{e}", parent=self.root)
        finally:
            self.root.destroy()

def main(root=None):
    if root is None:
        # 単体起動の場合
        root = tk.Tk()
        standalone = True
    else:
        # 親（実行.py）から呼ばれた場合
        root = tk.Toplevel(root)
        standalone = False

    # ウィンドウ設定
    root.deiconify() 
    root.lift()
    root.focus_force()
    
    app = FinalProcessorApp(root) 
    root.after(100, app.start_process)
    
    if standalone:
        root.mainloop()
    else:
        # 05が終わるまで親側を待機させる
        root.wait_window()


if __name__ == "__main__":
    main()