"""
送信ログの医療機関名セルに色がついている場合、名簿シートの同じ医療機関名の行の
「送信準備可否」列に「可能」を入力するツール（既に「可能」の行は上書きせず、
件数のみ完了メッセージで通知する）。
さらに、名簿シートで以下のいずれかに該当する行を非表示にする。
  ・「オプション」列が健康診断非表示条件.jsonに記載された文字列を含む
  ・「送信準備可否」列の値が「可能」ではない
  ・「送信日時」列にデータが入っている（送信済み）
その後、表示中の行だけを「部門1名」で部門ごとに分割してExcelファイルを作成し、
さらに各ファイルをPNG画像に変換する（いずれもDownloads\分割済み健康診断名簿に保存）。

使い方:
  このファイルをダブルクリック（または python 医療機関色付けツール.py）で実行。
  ダイアログが開くので、対象のExcelファイル（.xlsx）を選択する。
  処理後、同じファイルに上書き保存される。
  PNG変換にはExcel（pywin32経由）が必要。
"""

import copy
import datetime
import json
import math
import os
import sys
import unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter

try:
    import win32com.client as win32com_client
except ImportError:
    win32com_client = None

try:
    from PIL import Image
except ImportError:
    Image = None

LOG_SHEET_NAME = "送信ログ"
LOG_HEADER_NAME = "医療機関名"
LOG_SHEET_NAME_COLUMN = "F"  # 送信ログで各行に対応するシート名が入っている列
LOG_TRANSCRIBE_COLUMN = "G"  # 各シートのH1の値の転記先の列
LOG_TRANSCRIBE_HEADER_NAME = "メッセージ"  # 転記先列（G1）の見出し
LOG_SENT_DATE_COLUMN = "H"  # 医療機関ごとの送信（分割）日を記録する列
MEIBO_SHEET_NAME = "名簿"
MEIBO_INSTITUTION_HEADER_NAME = "医療機関名"  # 名簿シートで医療機関名が入っている列の見出し
OPTION_HEADER_NAME = "オプション"
READY_HEADER_NAME = "送信準備可否"  # この列の値がREADY_OK_VALUEでない行を非表示にする
READY_OK_VALUE = "可能"
HIDE_CONDITIONS_JSON_PATH = r"C:\Users\owner\Desktop\健康診断非表示条件.json"

# 名簿シートの列表示レイアウト
MEIBO_HIDE_COLUMN_RANGES = [("E", "L"), ("R", "AN")]
MEIBO_VISIBLE_COLUMNS = {"F"}  # 上記の非表示範囲内でも表示にする列
MEIBO_COLUMN_WIDTHS = {"O": 15, "P": 15}

# 部門ごとの分割ファイル出力設定
DEPARTMENT_HEADER_NAME = "部門1名"
SPLIT_OUTPUT_DIR = r"C:\Users\owner\Downloads\分割済み健康診断名簿"
INVALID_FILENAME_CHARS = '\\/:*?"<>|'
LOG_FILE_PATH = r"C:\Users\owner\Downloads\健康診断処理ログ.txt"
SPLIT_BORDER_THIN_SIDE = Side(style="thin")
SPLIT_BORDER_THICK_SIDE = Side(style="medium")  # ヘッダー下側・A列右側用（1段階太い罫線）
WRAP_LINE_HEIGHT = 15  # 折り返し表示1行あたりの高さ(pt)
OPTION_COLUMN_DEFAULT_WIDTH = 8.43  # オプション列の幅が未設定の場合に使うExcel標準幅
SPLIT_PRINT_MARGIN_LR = 1.0  # 分割ファイル印刷時の左右余白（インチ）。印刷で端が切れないようにする
MEIBO_SENT_DATETIME_COLUMN = "AF"  # 名簿シートで行ごとの送信日時を記録する列
MEIBO_SENT_DATETIME_HEADER = "送信日時"
SPLIT_HEADER_CENTER_TEXT = "健康診断予約のお知らせ"  # 分割ファイルの印刷ヘッダー中央に入れる文言
PNG_MARGIN_PIXELS = 30  # PNG化する際に画像の周囲へ付ける白い余白（px）。印刷時に端が切れないようにする


class ProgressDialog:
    """処理の進捗を表示する簡易ダイアログ。update()呼び出しごとに再描画する。"""

    def __init__(self, title: str = "処理中..."):
        self.root = tk.Tk()
        self.root.title(title)
        self.root.attributes("-topmost", True)
        self.root.resizable(False, False)
        self.label = tk.Label(self.root, text="", width=60, anchor="w")
        self.label.pack(padx=20, pady=(15, 5))
        self.bar = ttk.Progressbar(self.root, orient="horizontal", length=340, mode="determinate")
        self.bar.pack(padx=20, pady=(0, 15))
        self.root.update()

    def update(self, current: int, total: int, text: str = ""):
        total = max(total, 1)
        self.bar["maximum"] = total
        self.bar["value"] = current
        self.label["text"] = f"{text} ({current}/{total})" if text else f"{current}/{total}"
        self.root.update()

    def close(self):
        self.root.destroy()


def pick_file() -> str:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title="対象のExcelファイルを選択してください",
        initialdir=r"C:\Users\owner\Desktop",
        filetypes=[("Excel ファイル", "*.xlsx"), ("すべてのファイル", "*.*")],
    )
    root.destroy()
    return path


def find_header_column(ws, header_name: str, header_row: int = 1) -> int:
    for cell in ws[header_row]:
        if cell.value == header_name:
            return cell.column
    raise ValueError(f"シート「{ws.title}」に見出し「{header_name}」が見つかりません")


def mark_institutions_as_sent(ws_log, log_col: int, institution_names: set, target_date: datetime.date) -> int:
    """institution_names に含まれる医療機関名の行の送信ログH列に target_date を記録する。"""
    sent_col = column_index_from_string(LOG_SENT_DATE_COLUMN)
    marked = 0
    for row in range(2, ws_log.max_row + 1):
        name = ws_log.cell(row=row, column=log_col).value
        if not name or not isinstance(name, str):
            continue
        if name.strip() in institution_names:
            ws_log.cell(row=row, column=sent_col).value = target_date
            marked += 1
    return marked


def load_hide_conditions(path: str) -> list:
    """健康診断非表示条件.json を読み込む。JSON配列でも、1行1条件のテキストでも対応。"""
    with open(path, "r", encoding="utf-8-sig") as f:
        raw = f.read()
    try:
        data = json.loads(raw)
        if isinstance(data, list):
            return [str(v).strip() for v in data if str(v).strip()]
    except json.JSONDecodeError:
        pass
    return [line.strip() for line in raw.splitlines() if line.strip()]


def transcribe_sheet_h1_to_log(wb, ws_log, log_col: int) -> dict:
    """
    送信ログのG1の見出しを「メッセージ」に設定し、F列に記載された各シート名のシートの
    H1の値を、送信ログの対応する行のG列へ転記する（転記の向きは 各シート→送信ログ）。
    あわせて、医療機関名（log_col）ごとの転記済みメッセージも返す。
    """
    f_col = column_index_from_string(LOG_SHEET_NAME_COLUMN)
    g_col = column_index_from_string(LOG_TRANSCRIBE_COLUMN)

    ws_log.cell(row=1, column=g_col).value = LOG_TRANSCRIBE_HEADER_NAME

    row_sheet_names = {}
    for row in range(2, ws_log.max_row + 1):
        name = ws_log.cell(row=row, column=f_col).value
        if name and str(name).strip():
            row_sheet_names[row] = str(name).strip()

    target_names = set(row_sheet_names.values())
    missing_sheets = sorted(name for name in target_names if name not in wb.sheetnames)

    original_h1 = {
        name: wb[name]["H1"].value for name in target_names if name in wb.sheetnames
    }

    transcribed_rows = 0
    institution_messages = {}
    for row, sheet_name in row_sheet_names.items():
        if sheet_name not in original_h1:
            continue
        message = original_h1[sheet_name]
        ws_log.cell(row=row, column=g_col).value = message
        transcribed_rows += 1

        inst_name = ws_log.cell(row=row, column=log_col).value
        if inst_name and str(inst_name).strip():
            institution_messages[str(inst_name).strip()] = message

    return {
        "transcribed_rows": transcribed_rows,
        "missing_sheets": missing_sheets,
        "institution_messages": institution_messages,
    }


def mark_ready_from_log_colors(ws_log, log_col: int, ws_meibo, meibo_col: int, ready_col: int) -> dict:
    """
    送信ログの医療機関名セルに色がついている場合、その医療機関名と一致する
    名簿シートの行の送信準備可否列にREADY_OK_VALUE（「可能」）を入力する。
    既にREADY_OK_VALUEになっている行は上書きせず、件数だけ別途集計する
    （呼び出し側で注意メッセージとして表示するため）。
    """
    colored_names = set()
    for row in range(2, ws_log.max_row + 1):
        cell = ws_log.cell(row=row, column=log_col)
        name = cell.value
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        if cell.fill and cell.fill.patternType == "solid":
            colored_names.add(name)

    updated = 0
    already_ready = 0
    if not colored_names:
        return {"updated": updated, "already_ready": already_ready}

    for row in range(2, ws_meibo.max_row + 1):
        v = ws_meibo.cell(row=row, column=meibo_col).value
        if not (isinstance(v, str) and v.strip() in colored_names):
            continue
        ready_cell = ws_meibo.cell(row=row, column=ready_col)
        current = str(ready_cell.value).strip() if ready_cell.value else ""
        if current == READY_OK_VALUE:
            already_ready += 1
        else:
            ready_cell.value = READY_OK_VALUE
            updated += 1

    return {"updated": updated, "already_ready": already_ready}


def hide_rows(ws, option_col, ready_col: int, sent_col: int, conditions: list) -> dict:
    """
    以下のいずれかに該当する行を非表示にする。
      ・オプション列の値が conditions のいずれかを含む（option_colがNoneの場合は判定しない）
      ・送信準備可否列の値がREADY_OK_VALUE（「可能」）ではない
      ・送信日時列にデータが入っている（送信済み）
    """
    hidden_by_option = 0
    hidden_by_not_ready = 0
    hidden_by_sent = 0
    for row in range(2, ws.max_row + 1):
        option_hit = False
        if option_col is not None:
            v = ws.cell(row=row, column=option_col).value
            if v and any(cond in str(v) for cond in conditions):
                option_hit = True

        ready_value = ws.cell(row=row, column=ready_col).value
        not_ready_hit = str(ready_value).strip() != READY_OK_VALUE if ready_value else True

        sent_hit = bool(ws.cell(row=row, column=sent_col).value)

        # 過去の実行で非表示になった行が、条件を満たさなくなった後も
        # 非表示のまま残らないよう、毎回hiddenを明示的に設定し直す。
        ws.row_dimensions[row].hidden = option_hit or not_ready_hit or sent_hit
        if option_hit:
            hidden_by_option += 1
        if not_ready_hit:
            hidden_by_not_ready += 1
        if sent_hit:
            hidden_by_sent += 1

    return {
        "option": hidden_by_option,
        "not_ready": hidden_by_not_ready,
        "sent": hidden_by_sent,
    }


def last_visible_column(ws, max_col: int) -> int:
    """max_col以下で、非表示になっていない最後（一番右）の列番号を返す。"""
    for col in range(max_col, 0, -1):
        if not ws.column_dimensions[get_column_letter(col)].hidden:
            return col
    return max_col


def apply_column_layout(ws):
    """名簿シートの列の非表示・幅を設定する（F列は非表示範囲内でも表示のまま）。"""
    for start, end in MEIBO_HIDE_COLUMN_RANGES:
        for col in range(column_index_from_string(start), column_index_from_string(end) + 1):
            letter = get_column_letter(col)
            if letter in MEIBO_VISIBLE_COLUMNS:
                continue
            ws.column_dimensions[letter].hidden = True

    for letter in MEIBO_VISIBLE_COLUMNS:
        ws.column_dimensions[letter].hidden = False

    for letter, width in MEIBO_COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width


def _estimate_wrapped_line_count(text: str, column_width: float) -> int:
    """列幅（文字単位）に対して、全角文字を2、半角文字を1として折り返した場合の行数を見積もる。"""
    width_units = max(1, column_width)
    total_lines = 0
    for segment in str(text).split("\n"):
        char_width = sum(
            2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in segment
        )
        total_lines += max(1, math.ceil(char_width / width_units))
    return max(1, total_lines)


def adjust_option_row_heights(ws, option_col: int) -> int:
    """
    名簿シートのオプション列は列幅を変えず、テキストが列幅に収まりきらない行だけ
    折り返し表示（wrap_text）にした上で行の高さを広げる。
    """
    letter = get_column_letter(option_col)
    column_width = ws.column_dimensions[letter].width or OPTION_COLUMN_DEFAULT_WIDTH

    adjusted = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=option_col)
        text = cell.value
        if not text or not str(text).strip():
            continue

        line_count = _estimate_wrapped_line_count(str(text), column_width)
        if line_count <= 1:
            continue

        existing = cell.alignment
        cell.alignment = Alignment(
            horizontal=existing.horizontal,
            vertical=existing.vertical or "top",
            wrap_text=True,
        )
        needed_height = line_count * WRAP_LINE_HEIGHT
        if needed_height > (ws.row_dimensions[row].height or 0):
            ws.row_dimensions[row].height = needed_height
            adjusted += 1

    return adjusted


def sanitize_filename(name: str) -> str:
    for ch in INVALID_FILENAME_CHARS:
        name = name.replace(ch, "_")
    return name.strip() or "未設定"


def append_completion_log(source_path: str, message: str) -> None:
    """完了時のメッセージを、実行日時・対象ファイル付きでログファイルに追記する。"""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entry = f"===== {timestamp}  対象: {source_path} =====\n{message}\n\n"
    with open(LOG_FILE_PATH, "a", encoding="utf-8") as f:
        f.write(entry)


def copy_cell_style(src_cell, dst_cell):
    """
    セルの値・書式をコピーする。値が数式（例: =IF(ISNUMBER(SEARCH("若年健診",Q1526)),"若年健診",IF(J1526>=35,"一般健診","定期健診"))）の場合、
    そのままコピーすると移動先の行に合わないセルを参照してしまうため、
    Translatorで移動先の行・列に合わせて相対参照を補正してからコピーする。
    """
    value = src_cell.value
    if isinstance(value, str) and value.startswith("="):
        value = Translator(value, origin=src_cell.coordinate).translate_formula(dst_cell.coordinate)
    dst_cell.value = value
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)


def split_by_department(
    ws_meibo,
    dept_col: int,
    institution_col: int,
    output_dir: str,
    institution_messages: dict = None,
    progress_callback=None,
) -> dict:
    """
    名簿シートの表示中の行だけを「部門1名」で部門ごとに分割し、
    output_dir に部門名.xlsx として保存する。
    同じ部門内で医療機関名が複数ある場合は、部門名_医療機関名.xlsx として
    医療機関ごとに別ファイルに分ける。
    名簿シートのAO列（送信日時）に本日の日付が入っている行は、本日分割済みの
    人とみなして対象から除外する（行単位で判定するため、同じ医療機関が複数の
    部門・店舗にまたがっていても正しく判定できる）。分割に含めた行のうち、
    AO列が空欄の行にのみ現在日時を記録する（既に日時が入っている行は上書きしない）。
    institution_messages に医療機関名のメッセージがある場合、各ファイルの最終行の次の行に
    メッセージ全文を書き込む。
    """
    institution_messages = institution_messages or {}
    ao_col = column_index_from_string(MEIBO_SENT_DATETIME_COLUMN)
    ws_meibo.cell(row=1, column=ao_col).value = MEIBO_SENT_DATETIME_HEADER

    today = datetime.date.today()

    dept_groups = {}
    already_sent_row_count = 0
    for row in range(2, ws_meibo.max_row + 1):
        if ws_meibo.row_dimensions[row].hidden:
            continue
        dept = ws_meibo.cell(row=row, column=dept_col).value
        if dept is None or str(dept).strip() == "":
            continue
        dept = str(dept).strip()

        sent_value = ws_meibo.cell(row=row, column=ao_col).value
        sent_date = None
        if isinstance(sent_value, datetime.datetime):
            sent_date = sent_value.date()
        elif isinstance(sent_value, datetime.date):
            sent_date = sent_value
        if sent_date == today:
            already_sent_row_count += 1
            continue

        inst = ws_meibo.cell(row=row, column=institution_col).value
        inst = str(inst).strip() if inst else ""

        dept_groups.setdefault(dept, {}).setdefault(inst, []).append(row)

    os.makedirs(output_dir, exist_ok=True)

    max_col = ws_meibo.max_column
    total = sum(len(inst_groups) for inst_groups in dept_groups.values())
    done = 0
    created_files = []
    processed_institutions = set()
    missing_message_files = []
    for dept, inst_groups in dept_groups.items():
        split_by_institution = len(inst_groups) > 1
        for inst, rows in inst_groups.items():
            done += 1
            if progress_callback:
                progress_callback(done, total, f"{dept} / {inst}" if inst else dept)

            if split_by_institution:
                filename = f"{sanitize_filename(dept)}_{sanitize_filename(inst)}.xlsx"
            else:
                filename = sanitize_filename(dept) + ".xlsx"
            out_path = os.path.join(output_dir, filename)

            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = MEIBO_SHEET_NAME
            new_ws.page_margins.left = SPLIT_PRINT_MARGIN_LR
            new_ws.page_margins.right = SPLIT_PRINT_MARGIN_LR

            title_row = 1
            header_row = title_row + 1

            # 非表示列に紛れて日付が見えなくならないよう、実際に表示される最後の列に置く。
            last_visible_col = last_visible_column(ws_meibo, max_col)

            title_cell = new_ws.cell(row=title_row, column=1)
            title_cell.value = SPLIT_HEADER_CENTER_TEXT
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            if last_visible_col > 1:
                new_ws.merge_cells(
                    start_row=title_row, start_column=1, end_row=title_row, end_column=last_visible_col - 1
                )
            date_cell = new_ws.cell(row=title_row, column=last_visible_col)
            date_cell.value = datetime.date.today().strftime("%Y/%m/%d")
            date_cell.alignment = Alignment(horizontal="right", vertical="center")

            for col in range(1, max_col + 1):
                copy_cell_style(ws_meibo.cell(row=1, column=col), new_ws.cell(row=header_row, column=col))
                letter = get_column_letter(col)
                src_dim = ws_meibo.column_dimensions[letter]
                dst_dim = new_ws.column_dimensions[letter]
                dst_dim.width = src_dim.width
                dst_dim.hidden = src_dim.hidden

            last_row = header_row
            for new_row, src_row in enumerate(rows, start=header_row + 1):
                for col in range(1, max_col + 1):
                    copy_cell_style(
                        ws_meibo.cell(row=src_row, column=col),
                        new_ws.cell(row=new_row, column=col),
                    )
                src_height = ws_meibo.row_dimensions[src_row].height
                if src_height:
                    new_ws.row_dimensions[new_row].height = src_height
                last_row = new_row

            border_last_row = last_row

            message = institution_messages.get(inst) if inst else None
            if message:
                message_row = last_row + 1
                message_cell = new_ws.cell(row=message_row, column=1)
                message_cell.value = message
                message_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                if max_col > 1:
                    new_ws.merge_cells(
                        start_row=message_row, start_column=1, end_row=message_row, end_column=max_col
                    )
                line_count = str(message).count("\n") + 1
                new_ws.row_dimensions[message_row].height = line_count * WRAP_LINE_HEIGHT
                border_last_row = message_row

            for r in range(title_row, border_last_row + 1):
                for c in range(1, max_col + 1):
                    new_ws.cell(row=r, column=c).border = Border(
                        left=SPLIT_BORDER_THIN_SIDE,
                        right=SPLIT_BORDER_THICK_SIDE if c == 1 else SPLIT_BORDER_THIN_SIDE,
                        top=SPLIT_BORDER_THIN_SIDE,
                        bottom=SPLIT_BORDER_THICK_SIDE if r == header_row else SPLIT_BORDER_THIN_SIDE,
                    )

            new_wb.save(out_path)
            created_files.append(out_path)
            if inst:
                processed_institutions.add(inst)
            if not message:
                missing_message_files.append(os.path.basename(out_path))

            sent_now = datetime.datetime.now()
            for src_row in rows:
                sent_cell = ws_meibo.cell(row=src_row, column=ao_col)
                if not sent_cell.value:
                    sent_cell.value = sent_now

    return {
        "created_files": created_files,
        "processed_institutions": processed_institutions,
        "already_sent_row_count": already_sent_row_count,
        "missing_message_files": missing_message_files,
    }


def _add_png_margin(png_path: str, margin_px: int) -> None:
    """PNG画像の四辺に白い余白を付ける（未編集のまま印刷しても表の端が切れないように）。"""
    with Image.open(png_path) as im:
        im = im.convert("RGB")
        padded = Image.new("RGB", (im.width + margin_px * 2, im.height + margin_px * 2), "white")
        padded.paste(im, (margin_px, margin_px))
    padded.save(png_path, "PNG")


def _is_png_effectively_blank(png_path: str, threshold: float = 0.003) -> bool:
    """白（に近い）ピクセルの割合がthreshold以上なら、実質白紙の画像とみなす。"""
    with Image.open(png_path) as im:
        gray = im.convert("L")
        w, h = gray.size
        if w == 0 or h == 0:
            return True
        histogram = gray.histogram()
        non_white = sum(histogram[:250])
        return (non_white / (w * h)) < threshold


def _capture_range_to_png(ws, rng, png_path: str) -> None:
    """
    範囲をコピーしてChartに貼り付け、PNGとして書き出す。
    Appearance:=xlPrinter を優先し（xlScreenは画面外に置いたウィンドウの描画待ちで
    白紙になることがあるため）、xlPrinterが使えない環境（既定のプリンタ未設定など）では
    xlScreenにフォールバックする。書き出し後に画像が白紙でないか検証し、白紙であれば
    もう一方の方式で再試行する。
    """
    appearance_order = (2, 1)  # xlPrinter, xlScreen
    last_error = None
    for appearance in appearance_order:
        chart_obj = None
        try:
            rng.CopyPicture(appearance, 2)  # Format:=xlBitmap
            chart_obj = ws.ChartObjects().Add(0, 0, rng.Width, rng.Height)
            chart_obj.Chart.Paste()
            chart_obj.Chart.Export(png_path, "PNG")
        except Exception as e:
            last_error = e
            continue
        finally:
            if chart_obj is not None:
                chart_obj.Delete()

        if Image is None or not _is_png_effectively_blank(png_path):
            return
        last_error = RuntimeError("出力されたPNGが白紙でした")

    raise last_error or RuntimeError("PNGの書き出しに失敗しました")


def convert_files_to_png(xlsx_paths: list, progress_callback=None) -> dict:
    """
    Excelファイルの一覧を、それぞれ同じフォルダ・同じファイル名のPNGに変換する。
    （名簿シートに実際に表示されている範囲だけを画像化。非表示列・非表示行は写らない）
    PNGをそのまま（未編集で）印刷しても表の端が切れないよう、画像の四辺に余白を付ける。
    Excelの操作画面を出したくないため、ウィンドウを画面外に移動して処理する。
    """
    if win32com_client is None:
        raise RuntimeError("pywin32がインストールされていないため、PNG変換できません（pip install pywin32）。")

    created = []
    errors = []

    excel = win32com_client.DispatchEx("Excel.Application")
    excel.Visible = True
    excel.DisplayAlerts = False
    excel.Left = -3000
    excel.Top = -3000
    total = len(xlsx_paths)
    try:
        for i, xlsx_path in enumerate(xlsx_paths, start=1):
            if progress_callback:
                progress_callback(i, total, os.path.splitext(os.path.basename(xlsx_path))[0])
            wb = None
            try:
                wb = excel.Workbooks.Open(xlsx_path)
                ws = wb.Sheets(1)
                rng = ws.UsedRange
                # 折り返し行の高さを実際の描画に合わせて再計算する。
                # ただし結合セル（メッセージ行）はAutoFitが正しく高さを計算できず
                # 文字が見切れてしまうため、結合されていない行だけ対象にする。
                first_row = rng.Row
                last_row_num = first_row + rng.Rows.Count - 1
                for r in range(first_row, last_row_num + 1):
                    if not ws.Cells(r, 1).MergeCells:
                        ws.Rows(r).AutoFit()

                png_path = os.path.splitext(xlsx_path)[0] + ".png"
                _capture_range_to_png(ws, rng, png_path)
                if Image is not None:
                    _add_png_margin(png_path, PNG_MARGIN_PIXELS)
                created.append(png_path)
            except Exception as e:
                errors.append(f"{os.path.basename(xlsx_path)}: {e}")
            finally:
                if wb is not None:
                    wb.Close(False)
    finally:
        excel.Quit()

    if Image is None and created:
        errors.append(
            "Pillow(PIL)が未インストールのため、PNGへの余白追加はスキップされました（pip install pillow）。"
        )

    return {"created": created, "errors": errors}


def main():
    path = pick_file()
    if not path:
        return

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        messagebox.showerror("エラー", f"ファイルを開けませんでした。\n{e}")
        return

    if LOG_SHEET_NAME not in wb.sheetnames:
        messagebox.showerror("エラー", f"シート「{LOG_SHEET_NAME}」が見つかりません。")
        return
    if MEIBO_SHEET_NAME not in wb.sheetnames:
        messagebox.showerror("エラー", f"シート「{MEIBO_SHEET_NAME}」が見つかりません。")
        return

    ws_log = wb[LOG_SHEET_NAME]
    ws_meibo = wb[MEIBO_SHEET_NAME]

    log_col = find_header_column(ws_log, LOG_HEADER_NAME)
    sync_result = transcribe_sheet_h1_to_log(wb, ws_log, log_col)

    meibo_col = find_header_column(ws_meibo, MEIBO_INSTITUTION_HEADER_NAME)
    ready_col = find_header_column(ws_meibo, READY_HEADER_NAME)
    sent_col = column_index_from_string(MEIBO_SENT_DATETIME_COLUMN)

    # 送信ログで医療機関名に色がついていれば、名簿の送信準備可否に「可能」を入力する
    ready_from_color = mark_ready_from_log_colors(ws_log, log_col, ws_meibo, meibo_col, ready_col)

    # ・オプション列が非表示条件.jsonに一致する行
    # ・送信準備可否列の値が「可能」ではない行
    # ・送信日時列にデータが入っている行
    # のいずれかを非表示にする
    hide_error = None
    conditions = []
    try:
        conditions = load_hide_conditions(HIDE_CONDITIONS_JSON_PATH)
    except FileNotFoundError:
        hide_error = f"「{HIDE_CONDITIONS_JSON_PATH}」が見つからないため、オプション列の非表示条件はスキップしました（送信準備可否・送信日時による非表示は実行します）。"

    try:
        option_col = find_header_column(ws_meibo, OPTION_HEADER_NAME)
    except ValueError as e:
        option_col = None
        hide_error = (hide_error + "\n" if hide_error else "") + str(e)

    hidden_counts = hide_rows(ws_meibo, option_col, ready_col, sent_col, conditions)

    apply_column_layout(ws_meibo)

    option_rows_adjusted = 0
    if option_col is not None:
        option_rows_adjusted = adjust_option_row_heights(ws_meibo, option_col)

    try:
        wb.save(path)
    except PermissionError:
        messagebox.showerror(
            "エラー",
            "ファイルに保存できませんでした。\nExcelで開いている場合は閉じてから再実行してください。",
        )
        return

    today = datetime.date.today()

    split_error = None
    created_files = []
    already_sent_row_count = 0
    missing_message_files = []
    try:
        dept_col = find_header_column(ws_meibo, DEPARTMENT_HEADER_NAME)
        progress = ProgressDialog("部門ごとに分割中...")
        try:
            split_result = split_by_department(
                ws_meibo,
                dept_col,
                meibo_col,
                SPLIT_OUTPUT_DIR,
                institution_messages=sync_result["institution_messages"],
                progress_callback=progress.update,
            )
        finally:
            progress.close()
        created_files = split_result["created_files"]
        already_sent_row_count = split_result["already_sent_row_count"]
        missing_message_files = split_result["missing_message_files"]

        if split_result["processed_institutions"]:
            # 名簿のAO列（行単位の送信日時）はsplit_by_department内で既に記録済み。
            # 送信ログのH列は判定には使わず、参考記録として残す。
            mark_institutions_as_sent(ws_log, log_col, split_result["processed_institutions"], today)
            try:
                wb.save(path)
            except PermissionError:
                messagebox.showerror(
                    "エラー",
                    "送信履歴の保存に失敗しました。\nExcelで開いている場合は閉じてから再実行してください。",
                )
    except ValueError as e:
        split_error = str(e)

    png_result = {"created": [], "errors": []}
    png_error = None
    if created_files:
        try:
            progress = ProgressDialog("PNGに変換中...")
            try:
                png_result = convert_files_to_png(created_files, progress_callback=progress.update)
            finally:
                progress.close()
        except RuntimeError as e:
            png_error = str(e)

    msg = (
        f"各シートのH1から送信ログのG列へ転記した行数: {sync_result['transcribed_rows']}件\n"
        f"送信ログの色付けにより送信準備可否を「{READY_OK_VALUE}」にした行: {ready_from_color['updated']}件\n"
        f"オプション列の非表示条件に一致して非表示にした行: {hidden_counts['option']}件\n"
        f"送信準備可否が「{READY_OK_VALUE}」ではないため非表示にした行: {hidden_counts['not_ready']}件\n"
        f"送信日時にデータが入っているため非表示にした行: {hidden_counts['sent']}件\n"
        f"オプション列が収まらず行の高さを広げた行: {option_rows_adjusted}件\n"
        f"部門ごとに分割したファイル: {len(created_files)}件（{SPLIT_OUTPUT_DIR}）\n"
        f"PNGに変換したファイル: {len(png_result['created'])}件\n"
        f"本日送信済み（AO列）のためスキップした人数: {already_sent_row_count}件\n"
        f"メッセージが入らなかった分割ファイル: {len(missing_message_files)}件\n"
    )
    if ready_from_color["already_ready"]:
        msg += (
            f"\n\n[注意] 送信ログで色がついていますが、名簿側は既に「{READY_OK_VALUE}」だった行: "
            f"{ready_from_color['already_ready']}件"
        )
    if missing_message_files:
        msg += (
            "\n\n[注意] メッセージが入らなかった分割ファイルがあります:\n"
            + "\n".join(missing_message_files)
        )
    if sync_result["missing_sheets"]:
        msg += (
            "\n\n[注意] 送信ログのF列に記載されたシートが見つかりませんでした:\n"
            + "\n".join(sync_result["missing_sheets"])
        )
    if hide_error:
        msg += f"\n\n[注意] {hide_error}"
    if split_error:
        msg += f"\n\n[注意] 部門分割をスキップしました: {split_error}"
    if png_error:
        msg += f"\n\n[注意] PNG変換をスキップしました: {png_error}"
    if png_result["errors"]:
        msg += "\n\n[注意] PNG変換に失敗したファイル:\n" + "\n".join(png_result["errors"][:10])

    try:
        append_completion_log(path, msg)
    except OSError:
        pass

    messagebox.showinfo("完了", msg)


if __name__ == "__main__":
    main()
