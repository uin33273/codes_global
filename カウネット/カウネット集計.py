# このアプリは、前処理済みのCSVファイルとマスターファイル（Excel）を選択して、注文データを集計・加工し、最終的にExcelファイルとして保存するものです。
# 主な機能は以下の通りです：
# 1. CSVファイルを読み込み、注文データを結合する。
# 2. マスターファイルから店舗名のリストを取得し、お届け先部署名を変換して店番付き店舗名を作成する。
# 3. 注文番号ごとに合計金額を計算し、商品名を加工して代表的な商品名を作成する。
# 4. 最終的な集計結果をExcelファイルとして保存し、書式設定を行う。
# 代表商品名は金額最大のものを選ぶが、「少額受注配送料」になる場合は他の商品名を採用する。
#
# ■ お届け先部署名 → 店番付き店舗名 変換ルール
#
# 【除去・置換】（この順で処理）
#   「店」         → 削除
#   「グローバル」  → 削除
#   全角英数字     → 半角に統一（例：Ａ→A）
#   「前橋青葉」   → 「前橋青葉町」に置換
#
# 【プレフィックス変換】
#   メソッド（キッズの有無を問わず） → KM
#   キッズ（メソッドなし）           → K
#   パーク                           → P  （KP、KMPなど組み合わせ可）
#   サカフル                         → SF
#   ナーシングホーム                 → NH （ホームより優先）
#   ホーム（グローバルホーム）       → GH
#   ツリー                           → T
#
# 【特別扱い】
#   平出事務所     → そのまま「平出事務所」
#   お届け先空欄  → 「空欄」（紫色セル）
#   一致なし・複数一致 → 「未登録」（黄色セル）
#
# 【変換例】
#   グローバルキッズメソッド岩曽店       → KM岩曽    → 001KM岩曽
#   グローバルキッズパークせんげん台店   → KP せんげん台 → 001KPせんげん台
#   メソッド新栃木店                     → KM新栃木  → 157KM新栃木
#   グローバルホーム加須店               → GH加須    → 003GH加須
#   グローバルワークスA店                → ワークスA → 001ワークスA
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
import re
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- 1. 実際の処理を行う関数 ---
def process_order_data():
    root = tk.Tk()
    root.withdraw()

    downloads = str(Path.home() / "Downloads")
    messagebox.showinfo("手順1", "カウネット中間データファイルを選択してください")
    file_path = filedialog.askopenfilename(title="CSV選択", filetypes=[("CSV", "*.csv")], initialdir=downloads)
    if not file_path: return

    messagebox.showinfo("手順2", "table_HHHL共通店舗一覧mを選択してください")
    master_path = filedialog.askopenfilename(title="マスター選択", filetypes=[("Excel", "*.xlsx *.xls *.xlsm")])
    if not master_path: return

    try:
        # --- マスターの読み込み ---
        master_df = pd.read_excel(master_path, sheet_name='リスト', header=None)
        master_list = master_df.iloc[:, 0].dropna().astype(str).str.strip().tolist()

        # --- CSVの読み込み ---
        df = pd.read_csv(file_path, encoding='cp932')
        df.columns = df.columns.str.strip()

        # --- 合計金額の計算 ---
        df['税込小計'] = pd.to_numeric(df['税込小計'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        total_amounts = df.groupby('伝票番号')['税込小計'].sum().reset_index(name='合計金額')

        # --- 代表商品名の選定ロジック ---
        df_sorted = df.sort_values(['伝票番号', '税込小計'], ascending=[True, False])

        def get_representative_name(group):
            top_item = group.iloc[0]['商品名']
            if top_item == "少額受注配送料" and len(group) > 1:
                other_items = group[group['商品名'] != "少額受注配送料"]
                if not other_items.empty:
                    return other_items.iloc[0]['商品名']
            return top_item

        rep_names = df_sorted.groupby('伝票番号').apply(get_representative_name).reset_index(name='商品名')
        representative = df_sorted.drop_duplicates('伝票番号').copy()
        representative = representative.drop(columns=['商品名']).merge(rep_names, on='伝票番号')

        # --- 商品名の加工 ---
        order_counts = df.groupby('伝票番号').size().reset_index(name='商品数')
        representative = pd.merge(representative, order_counts, on='伝票番号')

        def format_product_name(row):
            name = str(row['商品名'])
            parts = re.split(r'[ 　]', name)
            if len(parts) >= 3:
                base_name = f"{parts[0]} {parts[1]}"
            elif len(parts) == 2:
                base_name = parts[0]
            else:
                base_name = name
            return f"{base_name}_他" if row['商品数'] > 1 else base_name

        representative['商品名'] = representative.apply(format_product_name, axis=1)

        # --- 変換店舗名の作成 ---
        def create_conv_name(target):
            if not isinstance(target, str): return ""
            # 平出事務所はそのまま返す
            if str(target).strip() == "平出事務所":
                return "平出事務所"
            name = target.replace("店", "").replace("グローバル", "").replace(" ", "").replace("　", "")
            name = name.replace("前橋青葉", "前橋青葉町")
            prefix = ""
            if "メソッド" in name: prefix += "KM"
            elif "キッズ" in name: prefix += "K"
            if "パーク" in name: prefix += "P"
            if "サカフル" in name: prefix += "SF"
            if "ナーシングホーム" in name: prefix += "NH"
            elif "ホーム" in name: prefix += "GH"
            if "ツリー" in name: prefix += "T"
            parts = re.split(r'キッズ|メソッド|パーク|サカフル|ナーシングホーム|ホーム|ツリー', name)
            sub_name = parts[-1] if len(parts) > 1 else name
            # 店名部分が空（prefixのみ）の場合は空欄扱いにする
            if sub_name == "":
                return ""
            return f"{prefix}{sub_name}"

        representative['変換店舗名'] = representative['お届け先部署名'].apply(create_conv_name)

        # --- マスターと照合 ---
        def find_full_name(row):
            dept = row['お届け先部署名']
            if not isinstance(dept, str) or dept.strip() == "":
                return "空欄"
            conv_name = row['変換店舗名']
            # 店名部分なし（グローバルキッズメソッドのみ等）も空欄扱い
            if not conv_name: return "空欄"
            # マスターは「番号+prefix+店名」形式（例: 001KM岩曽）
            # conv_nameは「prefix+店名」形式（例: KM岩曽）なので先頭の番号を除いて完全一致照合
            def normalize(s):
                import unicodedata
                s = unicodedata.normalize('NFKC', s)
                return s.replace("店", "").replace(" ", "").replace("　", "").strip()
            matches = [m for m in master_list if normalize(re.sub(r'^\d+', '', m)) == normalize(conv_name)]
            # 一致なし または 複数一致 → 未登録
            if len(matches) != 1:
                return "未登録"
            return matches[0]

        representative['店番付き店舗名'] = representative.apply(find_full_name, axis=1)
        result = pd.merge(representative, total_amounts, on='伝票番号')
        result['取引先'] = result.apply(lambda x: f"㈲コパン（㈱カウネット）{x['口座番号']}", axis=1)

        # --- 3重ソート ---
        result = result.sort_values(by=['ご登録電話番号', '出荷日', '伝票番号'], ascending=[True, True, True])

        # --- 列の並び替え ---
        cols = ["出荷日", "注文番号", "取引先", "商品名", "合計金額", "店番付き店舗名", "伝票番号", "ご登録電話番号", "お届け先部署名"]
        cols = [c for c in cols if c in result.columns]
        final_result = result.reindex(columns=cols)

        # --- 保存・書式設定 ---
        save_path = Path.home() / "Downloads" / "カウネットコピペデータ.xlsx"
        has_unregistered = False

        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            final_result.to_excel(writer, index=False, sheet_name='集計結果')
            ws = writer.sheets['集計結果']
            ws.freeze_panes = 'B2'
            tel_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")
            alert_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            tel_col_idx = cols.index("ご登録電話番号") + 1
            amount_col_idx = cols.index("合計金額") + 1
            master_col_idx = cols.index("店番付き店舗名") + 1

            prev_tel = None
            is_colored = False

            for row_idx in range(2, ws.max_row + 1):
                current_tel = ws.cell(row=row_idx, column=tel_col_idx).value
                if current_tel != prev_tel:
                    is_colored = not is_colored
                if is_colored:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = tel_fill

                master_cell = ws.cell(row=row_idx, column=master_col_idx)
                if master_cell.value == "未登録":
                    master_cell.fill = alert_fill
                    has_unregistered = True
                elif master_cell.value == "空欄":
                    master_cell.fill = PatternFill(start_color="E1BEE7", end_color="B3E5FC", fill_type="solid")
                ws.cell(row=row_idx, column=amount_col_idx).number_format = '#,##0'
                prev_tel = current_tel

            for i, col in enumerate(ws.columns, 1):
                column_letter = get_column_letter(i)
                if column_letter == 'C':
                    ws.column_dimensions[column_letter].width = 30
                else:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                val_len = len(str(cell.value).encode('shift_jis'))
                                if val_len > max_length: max_length = val_len
                        except: pass
                    ws.column_dimensions[column_letter].width = max_length + 2

        if has_unregistered:
            messagebox.showwarning("完了（警告あり）", "保存が完了しましたが、未登録の店舗があります。")
        else:
            messagebox.showinfo("完了", "保存が完了しました！")
        import os
        os.startfile(save_path)

    except PermissionError:
        messagebox.showwarning(
            "ファイルが開いています",
            "カウネットコピペデータ.xlsx が Excel で開いたままになっています。\n"
            "Excelを閉じてから再実行してください。"
        )
    except Exception as e:
        import traceback
        messagebox.showerror("エラー", f"エラーが発生しました:\n{traceback.format_exc()}")

# --- 2. 外部から呼ばれるためのmain関数 ---
def main():
    process_order_data()

# --- 3. 単体実行用 ---
if __name__ == "__main__":
    main()
