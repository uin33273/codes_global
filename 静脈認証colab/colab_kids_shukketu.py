# ============================================================
# Google Colab用 - 出退判定◯：◯以外を所属名別に分割
# 使い方:
# 1. このセルを実行 → ファイルをアップロード
# 2. 日付を選択（任意・未選択なら全期間）して「実行」ボタンを押す
# 3. 所属名ごとのExcelファイルZIP＋画像のみZIPをダウンロード
# ============================================================

from google.colab import files, auth
from google.auth import default
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, Color
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
import ipywidgets as widgets
from IPython.display import display
import os, re, zipfile, copy, datetime, io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import pandas as pd
from PIL import Image, ImageChops

# ── 対象シート名（この文字列を含むシートを必ず使用） ──────
SHEET_KEYWORD = 'KIDS店舗出退勤'

def _get_target_sheet(wb):
    """SHEET_KEYWORD を含むシートを取得。見つからなければエラー。"""
    matches = [s for s in wb.sheetnames if SHEET_KEYWORD in s]
    if not matches:
        raise ValueError(
            f"'{SHEET_KEYWORD}' を含むシートが見つかりません。"
            f"シート一覧: {wb.sheetnames}"
        )
    return wb[matches[0]]

# ── 店舗別エラー件数の推移を蓄積するExcelファイル ──────────
# 保存先はこのDriveフォルダ内（ファイルが無ければ新規作成する）
TREND_FOLDER_ID  = '1epPSIGZy1OluWGQv6O3kh_X96FdVKnmt'
TREND_FILE_NAME  = 'KIDS店舗出退勤_エラー推移.xlsx'
TREND_SHEET_NAME = 'エラー推移'
TREND_HEADER     = ['日付', '所属名', 'エラー件数']
PIVOT_SHEET_NAME = '集計'
CHART_SHEET_NAME = 'グラフ'
TOP_N_STORES = 10
_XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

def _rebuild_trend_chart(wb):
    """エラー推移シートの生データから 日付×所属名 のピボット表を作り、
    合計エラー件数が多い上位 TOP_N_STORES 店舗だけを折れ線グラフに描画する（生データ自体は全店舗を保持）"""
    ws_src = wb[TREND_SHEET_NAME]
    rows = [r for r in ws_src.iter_rows(min_row=2, values_only=True) if r[0] is not None]

    dates = sorted({r[0] for r in rows})

    totals_by_store = {}
    for r in rows:
        totals_by_store[r[1]] = totals_by_store.get(r[1], 0) + (r[2] or 0)
    stores = sorted(totals_by_store, key=lambda s: totals_by_store[s], reverse=True)[:TOP_N_STORES]
    store_set = set(stores)

    value_map = {(r[0], r[1]): r[2] for r in rows if r[1] in store_set}

    # 日付ごとの合計（全店舗対象、TOP_N_STORESの絞り込みとは無関係）
    total_by_date = {}
    for r in rows:
        total_by_date[r[0]] = total_by_date.get(r[0], 0) + (r[2] or 0)

    TOTAL_LABEL = '日々のエラー合計'

    if PIVOT_SHEET_NAME in wb.sheetnames:
        del wb[PIVOT_SHEET_NAME]
    ws_pivot = wb.create_sheet(PIVOT_SHEET_NAME)
    ws_pivot.append(['日付'] + stores + [TOTAL_LABEL])
    for d in dates:
        ws_pivot.append([d] + [value_map.get((d, s)) for s in stores] + [total_by_date.get(d)])

    if CHART_SHEET_NAME in wb.sheetnames:
        del wb[CHART_SHEET_NAME]
    ws_chart = wb.create_sheet(CHART_SHEET_NAME)

    if not dates or not stores:
        return

    max_row = 1 + len(dates)
    store_max_col = 1 + len(stores)
    total_col = store_max_col + 1
    cats = Reference(ws_pivot, min_col=1, max_col=1, min_row=2, max_row=max_row)

    # ── 店舗別（左軸）── 合計を混ぜないことで店舗間の増減が見やすくなる
    chart = LineChart()
    chart.title = f'店舗別エラー件数の推移（上位{TOP_N_STORES}店舗）'
    chart.x_axis.title = '日付'
    chart.y_axis.title = 'エラー件数（店舗別）'
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.style = 2
    chart.width = 28
    chart.height = 14

    data_stores = Reference(ws_pivot, min_col=2, max_col=store_max_col, min_row=1, max_row=max_row)
    chart.add_data(data_stores, titles_from_data=True)
    chart.set_categories(cats)
    for s in chart.series:
        s.smooth = False

    # ── 日々のエラー合計（右軸）── 別スケールにして店舗別の折れ線を潰さないようにする
    chart_total = LineChart()
    data_total = Reference(ws_pivot, min_col=total_col, max_col=total_col, min_row=1, max_row=max_row)
    chart_total.add_data(data_total, titles_from_data=True)
    chart_total.y_axis.axId = 200
    chart_total.y_axis.title = 'エラー件数（日々の合計）'
    chart_total.y_axis.delete = False
    # 右軸として表示するため、x軸の最大側で交差させる
    chart_total.y_axis.crosses = 'max'
    for s in chart_total.series:
        s.smooth = False
        s.graphicalProperties.line.width = 28000  # EMU単位（約2.2pt）
        s.graphicalProperties.line.solidFill = 'FF0000'
        s.marker.symbol = 'square'
        s.marker.size = 7
        s.marker.graphicalProperties.solidFill = 'FF0000'
        s.marker.graphicalProperties.line.solidFill = 'FF0000'
        # ポイント付近に件数の数値を表示
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.showLegendKey = False
        s.dLbls.showCatName = False
        s.dLbls.showSerName = False
        s.dLbls.showPercent = False
        s.dLbls.showBubbleSize = False
        s.dLbls.dLblPos = 't'

    chart += chart_total
    ws_chart.add_chart(chart, 'B2')

def _find_trend_file_id(drive_service):
    query = (
        f"'{TREND_FOLDER_ID}' in parents and "
        f"name = '{TREND_FILE_NAME}' and trashed = false"
    )
    found = drive_service.files().list(q=query, fields='files(id, name)').execute().get('files', [])
    return found[0]['id'] if found else None

def _download_trend_workbook(drive_service, file_id):
    request = drive_service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return load_workbook(buf)

def _upload_trend_workbook(drive_service, wb, file_id):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    media = MediaIoBaseUpload(buf, mimetype=_XLSX_MIME, resumable=False)
    if file_id:
        drive_service.files().update(fileId=file_id, media_body=media).execute()
    else:
        metadata = {'name': TREND_FILE_NAME, 'parents': [TREND_FOLDER_ID]}
        drive_service.files().create(body=metadata, media_body=media, fields='id').execute()

def _update_trend_sheet(counts_by_date_store):
    """{(日付, 所属名): 件数} を、指定Driveフォルダ内のExcelファイルに蓄積（既存の同じ日付×所属名は何もせずスキップ）"""
    auth.authenticate_user()
    creds, _ = default()
    drive_service = build('drive', 'v3', credentials=creds)

    file_id = _find_trend_file_id(drive_service)
    if file_id:
        wb = _download_trend_workbook(drive_service, file_id)
        if TREND_SHEET_NAME in wb.sheetnames:
            ws = wb[TREND_SHEET_NAME]
        else:
            ws = wb.create_sheet(TREND_SHEET_NAME)
            ws.append(TREND_HEADER)
    else:
        print(f"'{TREND_FILE_NAME}' がフォルダ内に見つからないため新規作成します。")
        wb = Workbook()
        ws = wb.active
        ws.title = TREND_SHEET_NAME
        ws.append(TREND_HEADER)

    existing_keys = {
        (str(row[0].value), str(row[1].value))
        for row in ws.iter_rows(min_row=2)
        if row[0].value is not None
    }

    appended, skipped = 0, 0
    for (date_str, store), count in counts_by_date_store.items():
        if (date_str, store) in existing_keys:
            skipped += 1
        else:
            ws.append([date_str, store, count])
            appended += 1

    _rebuild_trend_chart(wb)
    _upload_trend_workbook(drive_service, wb, file_id)
    print(f"集計Excelに反映: 追加{appended}件 / 既存のためスキップ{skipped}件")

# ── アップロード ──────────────────────────────────────────
uploaded = files.upload()
src = list(uploaded.keys())[0]
print(f"アップロード完了: {src}")

# ── 日本語フォント設定（Colab環境） ──────────────────────
!apt-get -qq install -y fonts-noto-cjk > /dev/null 2>&1
fm.fontManager.__init__()
_jp_fonts = [f.name for f in fm.fontManager.ttflist if 'Noto' in f.name and 'CJK' in f.name]
_font_name = _jp_fonts[0] if _jp_fonts else 'DejaVu Sans'
plt.rcParams['font.family'] = _font_name

# ── ファイルから日付一覧を取得 ────────────────────────────
_wb_tmp = load_workbook(src, read_only=True)
_ws_tmp = _get_target_sheet(_wb_tmp)
print(f"使用シート: {_ws_tmp.title}")

# ヘッダー行からキー列のインデックスを自動検出
_header_row = next(_ws_tmp.iter_rows(min_row=1, max_row=1, values_only=True), None)
_headers = [str(v).strip() if v is not None else '' for v in _header_row]

def _find_col(candidates):
    """候補名リストのいずれかに部分一致する列インデックスを返す"""
    for i, h in enumerate(_headers):
        for c in candidates:
            if c in h:
                return i
    return None

_col_date  = _find_col(['日時', '日付', 'date', 'Date'])
_col_store = _find_col(['所属名', '店舗名', '部署名'])
_col_judge = _find_col(['出退判定', '判定'])

print(f"検出列 → 日付:{_col_date}列目({_headers[_col_date] if _col_date is not None else 'なし'})  "
      f"所属名:{_col_store}列目({_headers[_col_store] if _col_store is not None else 'なし'})  "
      f"判定:{_col_judge}列目({_headers[_col_judge] if _col_judge is not None else 'なし'})")

if _col_date is None:
    print("⚠ 日付列が見つかりませんでした。ヘッダー名を確認してください。")
    print(f"  全ヘッダー: {_headers}")
    _dates = []
else:
    _dates = sorted(set(
        str(row[_col_date])[:10] for row in _ws_tmp.iter_rows(min_row=2, values_only=True)
        if row[_col_date] and re.match(r'\d{4}-\d{2}-\d{2}', str(row[_col_date]))
    ))
_wb_tmp.close()

if not _dates:
    print("⚠ 日付データが見つかりませんでした。")
else:
    print(f"ファイル内の日付: {_dates[0]} ～ {_dates[-1]}  ({len(_dates)}日分)")

# ── 日付選択UI ────────────────────────────────────────────
_ALL = '（指定なし・全期間）'
_opts = [_ALL] + _dates

_mode = widgets.ToggleButtons(
    options=[('単日選択', 'single'), ('範囲指定', 'range')],
    description='モード:',
    button_style='info',
)
_date_single = widgets.Dropdown(
    options=_opts, value=_ALL, description='日付:',
    layout=widgets.Layout(width='320px')
)
_date_from = widgets.Dropdown(
    options=_opts, value=_ALL, description='開始日:',
    layout=widgets.Layout(width='320px')
)
_date_to = widgets.Dropdown(
    options=_opts, value=_ALL, description='終了日:',
    layout=widgets.Layout(width='320px')
)
_single_box = widgets.VBox([_date_single])
_range_box  = widgets.VBox([_date_from, _date_to])
_content    = widgets.VBox([_single_box])
_btn = widgets.Button(
    description='実行',
    button_style='success',
    layout=widgets.Layout(width='160px', margin='10px 0 0 0')
)
display(widgets.VBox([_mode, _content, _btn]))

def _on_mode_change(change):
    _content.children = [_single_box] if change['new'] == 'single' else [_range_box]
_mode.observe(_on_mode_change, names='value')

# ── スタイル設定 ──────────────────────────────────────────
MATCH_OK        = '◯ ： ◯'
EXCLUDE_HEADERS = {'従業員コード', '所定時間', '残業時間', '所属コード'}

m      = re.search(r'(出退勤.*)', src)
suffix = m.group(1) if m else src

wb_src = load_workbook(src)
ws_src = _get_target_sheet(wb_src)

all_visible  = [c for c in range(1, ws_src.max_column + 1)
                if not (ws_src.column_dimensions.get(get_column_letter(c)) and
                        ws_src.column_dimensions[get_column_letter(c)].hidden)]
visible_cols = [c for c in all_visible
                if (ws_src.cell(1, c).value or '').replace('\n', '').strip() not in EXCLUDE_HEADERS]

thin        = Side(style='thin')
border      = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_FILL = PatternFill(fill_type='solid', fgColor=Color(theme=7, tint=0.7999816888943144))

def copy_cell_style(src_cell, dst_cell, is_header=False):
    if is_header:
        dst_cell.fill = HEADER_FILL
    elif src_cell.fill and src_cell.fill.fill_type not in (None, 'none'):
        dst_cell.fill = copy.copy(src_cell.fill)
    src_bold = src_cell.font.bold if src_cell.font else False
    dst_cell.font = Font(name='游ゴシック', size=11, bold=src_bold if is_header else False)
    if src_cell.alignment:
        dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.border = border
    if isinstance(src_cell.value, datetime.time):
        dst_cell.number_format = 'h:mm'

# ── PNG生成ヘルパー ───────────────────────────────────────
MAX_ROWS_PER_IMAGE = 50

def save_table_as_png(df, out_path, title=''):
    """DataFrameをmatplotlibのテーブルとしてPNG保存。行数が多い場合は分割。"""
    n_cols = len(df.columns)
    col_labels = list(df.columns)
    all_rows = df.values.tolist()
    chunks = [all_rows[i:i+MAX_ROWS_PER_IMAGE] for i in range(0, max(len(all_rows), 1), MAX_ROWS_PER_IMAGE)]
    paths = []
    for idx, chunk in enumerate(chunks):
        n_rows = len(chunk)
        fig_h = max(1.2, 0.45 * n_rows + 0.8)
        fig_w = max(6, 1.5 * n_cols)
        fig, ax = plt.subplots(figsize=(fig_w, fig_h))
        ax.axis('off')
        tbl = ax.table(
            cellText=chunk,
            colLabels=col_labels,
            loc='center',
            cellLoc='center',
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(9)
        tbl.auto_set_column_width(col=list(range(n_cols)))
        for (r, c), cell in tbl.get_celld().items():
            if r == 0:
                cell.set_facecolor('#BDD7EE')
                cell.set_height(cell.get_height() * 2)
                cell.set_text_props(fontweight='bold', wrap=True)
            else:
                cell.set_facecolor('#FFFFFF' if r % 2 == 1 else '#F2F2F2')
            cell.set_edgecolor('#AAAAAA')

        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=150, bbox_inches='tight', pad_inches=0)
        plt.close(fig)
        buf.seek(0)
        img = Image.open(buf).convert('RGB')
        bg   = Image.new('RGB', img.size, (255, 255, 255))
        diff = ImageChops.difference(img, bg)
        bbox = diff.getbbox()
        if bbox:
            img = img.crop(bbox)

        p = out_path if len(chunks) == 1 else out_path.replace('.png', f'_{idx+1}.png')
        img.save(p, dpi=(150, 150))
        paths.append(p)
    return paths

def excel_to_df(xlsx_path):
    """openpyxlで読み込んでDataFrameに変換（time型を文字列化）"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    headers = [str(v) if v is not None else '' for v in rows[0]]
    data = []
    for row in rows[1:]:
        data.append([
            v.strftime('%H:%M') if isinstance(v, datetime.time)
            else (str(v) if v is not None else '')
            for v in row
        ])
    return pd.DataFrame(data, columns=headers)

# ── 実行ボタン処理 ────────────────────────────────────────
def _run(b):
    # 列インデックスの確認
    if _col_date is None or _col_store is None or _col_judge is None:
        print("⚠ 必要な列（日付・所属名・判定）が検出できていません。処理を中止します。")
        return
    if not _dates:
        print("⚠ 日付データが読み込めていません。ファイルを確認してください。")
        return

    # 日付フィルタ範囲を決定
    if _mode.value == 'single':
        sel = _date_single.value
        date_from = date_to = None if sel == _ALL else sel
    else:
        df = _date_from.value;  dt = _date_to.value
        date_from = None if df == _ALL else df
        date_to   = None if dt == _ALL else dt

    if date_from is None and date_to is None:
        label = '全期間'
    elif date_from == date_to:
        label = date_from
    else:
        label = f"{date_from or _dates[0]}～{date_to or _dates[-1]}"
    print(f"\n対象期間: {label}")

    # フィルタ条件でグループ化（行番号保持）
    data_by_store = {}
    counts_by_date_store = {}
    for src_row in ws_src.iter_rows(min_row=2):
        store = src_row[_col_store].value
        date  = src_row[_col_date].value
        judge = src_row[_col_judge].value
        if store is None or date is None:
            continue
        date_str = str(date)[:10]
        if date_from and date_str < date_from:
            continue
        if date_to and date_str > date_to:
            continue
        if str(judge) == MATCH_OK:
            continue
        store = str(store).strip()
        if store not in data_by_store:
            data_by_store[store] = []
        data_by_store[store].append(src_row[0].row)
        key = (date_str, store)
        counts_by_date_store[key] = counts_by_date_store.get(key, 0) + 1

    total = sum(len(v) for v in data_by_store.values())
    print(f"抽出行数: {total}行 / 所属名: {len(data_by_store)}店舗")
    if not data_by_store:
        print("⚠ 条件に一致するデータがありませんでした。")
        return

    # 各行の日時データに基づき、日付×所属名ごとの件数（分割ファイルに入る件数）を集計シートに反映
    try:
        _update_trend_sheet(counts_by_date_store)
    except Exception as e:
        print(f"⚠ 集計シートへの書き込みに失敗しました: {e}")

    # 出力フォルダ初期化
    out_dir = 'output_filtered'
    img_dir = 'output_images'
    for d in [out_dir, img_dir]:
        os.makedirs(d, exist_ok=True)
        for f in os.listdir(d):
            os.remove(os.path.join(d, f))

    all_png_paths = []

    # 所属名ごとに1ファイル作成
    for store, src_rows in data_by_store.items():
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = ws_src.title
        wb_out.loaded_theme = wb_src.loaded_theme

        for out_c, orig_c in enumerate(visible_cols, start=1):
            sc = ws_src.cell(1, orig_c); dc = ws_out.cell(1, out_c)
            dc.value = sc.value
            copy_cell_style(sc, dc, is_header=(out_c != len(visible_cols)))

        for out_c, orig_c in enumerate(visible_cols, start=1):
            ol = get_column_letter(orig_c); dl = get_column_letter(out_c)
            if ol in ws_src.column_dimensions:
                ws_out.column_dimensions[dl].width = ws_src.column_dimensions[ol].width

        for out_r, src_r in enumerate(src_rows, start=2):
            for out_c, orig_c in enumerate(visible_cols, start=1):
                sc = ws_src.cell(src_r, orig_c); dc = ws_out.cell(out_r, out_c)
                dc.value = sc.value
                copy_cell_style(sc, dc)

        safe     = store.translate(str.maketrans('/\\*?:"<>|', '／＼＊？："＜＞｜'))
        out_name = f"{safe}_{suffix}"
        xlsx_path = os.path.join(out_dir, out_name)
        wb_out.save(xlsx_path)
        print(f"  保存: {out_name} ({len(src_rows)}行)", end='')

        # PNG生成
        try:
            df_img = excel_to_df(xlsx_path)
            df_img = df_img.iloc[:, :-1]   # 右端の凡例列を除外
            png_name = out_name.replace('.xlsx', '.png')
            png_path = os.path.join(img_dir, png_name)
            saved_pngs = save_table_as_png(df_img, png_path, title=safe)
            all_png_paths.extend(saved_pngs)
            print(f"  → 画像: {[os.path.basename(p) for p in saved_pngs]}")
        except Exception as e:
            print(f"  → 画像生成エラー: {e}")

    zip_label = label.replace('～', '_')

    # Excel ZIP
    zip_excel = src.replace('.xlsx', f'_{zip_label}_分割.zip')
    with zipfile.ZipFile(zip_excel, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname in os.listdir(out_dir):
            zf.write(os.path.join(out_dir, fname), fname)
    print(f"\nExcel ZIP: {zip_excel}")
    files.download(zip_excel)

    # 画像専用 ZIP
    if all_png_paths:
        zip_images = src.replace('.xlsx', f'_{zip_label}_画像.zip')
        with zipfile.ZipFile(zip_images, 'w', zipfile.ZIP_DEFLATED) as zf:
            for p in all_png_paths:
                zf.write(p, os.path.basename(p))
        print(f"画像 ZIP: {zip_images}")
        files.download(zip_images)

    print("ダウンロード開始")
    print("更新終了しました")

_btn.on_click(_run)